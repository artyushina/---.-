import sys
import psycopg2
import re
from PyQt6.QtWidgets import (QApplication, QMainWindow, QTableWidgetItem,
                             QHeaderView, QMessageBox, QDialog, QVBoxLayout,
                             QLabel, QLineEdit, QDialogButtonBox, QComboBox,
                             QHBoxLayout, QPushButton, QTableWidget, QWidget,
                             QToolBar, QAbstractItemView, QCompleter, QScrollArea,
                             QCheckBox, QGroupBox, QSizePolicy, QListWidget, QListWidgetItem,
                             QMenu, QFileDialog, QFrame)
from PyQt6.QtGui import QAction, QColor, QFont
from PyQt6.QtCore import Qt, QSettings
from PyQt6.QtGui import QRegularExpressionValidator
from PyQt6.QtCore import QRegularExpression, QEvent
from config import DB_CONFIG
from datetime import datetime, date
import os
from MainFormlayout_11week import Ui_MainWindow
from PyQt6.QtCore import QTimer

# Импорты для экспорта отчетов
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class DateValidator:
    """Класс для валидации и форматирования дат"""

    @staticmethod
    def parse_date(date_string):
        """Парсит дату из строки в объект datetime"""
        if not date_string or not date_string.strip():
            return None

        formats = [
            '%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y',
            '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
            '%d.%m.%y', '%d/%m/%y', '%d-%m-%y'
        ]

        for fmt in formats:
            try:
                return datetime.strptime(date_string.strip(), fmt)
            except ValueError:
                continue
        return None

    @staticmethod
    def format_date_for_display(date_string):
        """Форматирует дату для отображения (dd.mm.yyyy)"""
        if not date_string:
            return ""

        date_obj = DateValidator.parse_date(date_string)
        if date_obj:
            return date_obj.strftime('%d.%m.%Y')
        return date_string

    @staticmethod
    def format_date_for_db(date_string):
        """Форматирует дату для базы данных (yyyy-mm-dd)"""
        if not date_string:
            return None

        date_obj = DateValidator.parse_date(date_string)
        if date_obj:
            return date_obj.strftime('%Y-%m-%d')
        return date_string

    @staticmethod
    def get_format_examples():
        """Возвращает примеры поддерживаемых форматов дат"""
        return "dd.mm.yyyy, dd/mm/yyyy, dd-mm-yyyy, yyyy-mm-dd"


class EditDialog(QDialog):
    """Универсальное диалоговое окно для других таблиц"""

    def __init__(self, table_name, columns, data=None, parent=None, display_names=None, date_columns=None,
                 is_edit=False, numeric_columns=None):
        super().__init__(parent)
        self.table_name = table_name
        self.columns = columns
        self.data = data
        self.display_names = display_names or {}
        self.date_columns = date_columns or []
        self.is_edit = is_edit
        self.numeric_columns = numeric_columns or []
        self.setup_ui()

    def setup_ui(self):
        # Устанавливаем заголовок в зависимости от режима и таблицы
        table_display_name = self.display_names.get('_table_name', self.table_name)
        if self.is_edit:
            self.setWindowTitle(f"Редактирование записи: {table_display_name}")
        else:
            self.setWindowTitle(f"Добавление записи: {table_display_name}")

        self.layout = QVBoxLayout()

        self.fields = {}
        for i, column in enumerate(self.columns):
            # Прячем технические идентификаторы
            if column.lower() == 'id':
                hidden_field = QLineEdit(self)
                if self.data and i < len(self.data):
                    field_value = self.data[i]
                    hidden_field.setText(str(field_value) if field_value is not None else "")
                else:
                    hidden_field.setText("")
                hidden_field.setReadOnly(True)
                hidden_field.hide()
                self.fields[column] = hidden_field
                continue

            label_text = self.display_names.get(column, column)
            label = QLabel(label_text)
            label.setStyleSheet("font-weight: bold; color: #2c3e50;")
            self.layout.addWidget(label)

            field = QLineEdit()
            if self.data and i < len(self.data):
                field_value = self.data[i]
                if field_value is None:
                    field_value = ""
                elif column in self.date_columns:
                    formatted_date = DateValidator.format_date_for_display(str(field_value))
                    field.setText(formatted_date if formatted_date else str(field_value))
                else:
                    field.setText(str(field_value))
            else:
                field.setText("")

            # Добавляем валидацию для числовых полей
            if column in self.numeric_columns:
                field.setValidator(QRegularExpressionValidator(QRegularExpression("[0-9]*")))

            field.setStyleSheet("""
                QLineEdit {
                    padding: 8px;
                    border: 1px solid #bdc3c7;
                    border-radius: 4px;
                    background-color: white;
                    margin-bottom: 10px;
                }
                QLineEdit:focus {
                    border-color: #3498db;
                }
                QLineEdit:read-only {
                    background-color: #ecf0f1;
                    color: #7f8c8d;
                }
            """)

            self.fields[column] = field
            self.layout.addWidget(field)

        if self.date_columns:
            hint_label = QLabel(f"Поддерживаемые форматы дат: {DateValidator.get_format_examples()}")
            hint_label.setStyleSheet("color: #7f8c8d; font-size: 10px; margin-bottom: 10px;")
            self.layout.addWidget(hint_label)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                                   QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.validate_and_accept)
        buttons.rejected.connect(self.reject)

        # Изменяем текст кнопок на русский
        ok_button = buttons.button(QDialogButtonBox.StandardButton.Ok)
        ok_button.setText("ОК")
        cancel_button = buttons.button(QDialogButtonBox.StandardButton.Cancel)
        cancel_button.setText("Отмена")

        # Стилизация кнопок диалога
        for button in buttons.buttons():
            if buttons.buttonRole(button) == QDialogButtonBox.ButtonRole.AcceptRole:
                button.setStyleSheet("""
                    QPushButton {
                        background-color: #27ae60;
                        color: white;
                        border: none;
                        padding: 8px 16px;
                        border-radius: 4px;
                        font-weight: bold;
                        min-width: 80px;
                    }
                    QPushButton:hover {
                        background-color: #219a52;
                    }
                    QPushButton:pressed {
                        background-color: #1e8449;
                    }
                """)
            else:
                button.setStyleSheet("""
                    QPushButton {
                        background-color: #95a5a6;
                        color: white;
                        border: none;
                        padding: 8px 16px;
                        border-radius: 4px;
                        font-weight: bold;
                        min-width: 80px;
                    }
                    QPushButton:hover {
                        background-color: #7f8c8d;
                    }
                    QPushButton:pressed {
                        background-color: #707b7c;
                    }
                """)

        self.layout.addWidget(buttons)

        self.setLayout(self.layout)

    def validate_and_accept(self):
        """Проверяет валидность данных перед принятием"""
        for column in self.date_columns:
            if column in self.fields:
                date_value = self.fields[column].text().strip()
                if date_value:
                    if not DateValidator.parse_date(date_value):
                        QMessageBox.warning(
                            self,
                            "Неверный формат даты",
                            f"Поле '{self.display_names.get(column, column)}' содержит неверный формат даты.\n\n"
                            f"Поддерживаемые форматы:\n{DateValidator.get_format_examples()}"
                        )
                        return
        self.accept()

    def get_data(self):
        """Получить данные из полей ввода"""
        result = []
        for column in self.columns:
            if column in self.fields:
                value = self.fields[column].text().strip()
                if column in self.date_columns:
                    if value:
                        db_date = DateValidator.format_date_for_db(value)
                        result.append(db_date if db_date else value)
                    else:
                        result.append(None)
                else:
                    # Для текстовых полей, если значение пустое, устанавливаем NULL
                    if not value:
                        result.append(None)
                    else:
                        result.append(value)
            else:
                result.append(None)
        return result


class MultiSelectComboBox(QWidget):
    """Виджет с выпадающим списком и автодополнением для множественного выбора"""

    def __init__(self, items=None, parent=None):
        super().__init__(parent)
        self.items = items or []
        self.selected_items = []
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        self.setLayout(layout)

        # ComboBox с возможностью редактирования
        self.combo_box = QComboBox()
        self.combo_box.setEditable(True)
        self.combo_box.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.combo_box.addItems([""] + self.items)

        # Настраиваем автодополнение
        completer = QCompleter(self.items)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        self.combo_box.setCompleter(completer)

        # Выбранные значения

        self.selected_list = QListWidget()
        self.selected_list.setMaximumHeight(100)
        self.selected_list.setSelectionMode(QListWidget.SelectionMode.SingleSelection)
        self.selected_list.setStyleSheet("""
            QListWidget {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
            }
            QListWidget::item {
                padding: 5px;
                border-bottom: 1px solid #ecf0f1;
            }
            QListWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)
        self.selected_list.itemDoubleClicked.connect(self.remove_selected_item)
        layout.addWidget(self.selected_list)


        # Кнопки управления
        button_layout = QHBoxLayout()
        self.add_button = QPushButton("Добавить")
        self.add_button.clicked.connect(self.add_item)
        self.add_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        self.clear_button = QPushButton("Очистить все")
        self.clear_button.clicked.connect(self.clear_selection)
        self.clear_button.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        button_layout.addWidget(self.add_button)
        button_layout.addWidget(self.clear_button)
        layout.addWidget(QLabel("Введите значение и нажмите 'Добавить':"))
        layout.addWidget(self.combo_box)
        layout.addLayout(button_layout)
        layout.addWidget(QLabel("Выбранные значения (двойной клик для удаления):"))
        layout.addWidget(self.selected_list)

    def add_item(self):
        current_text = self.combo_box.currentText().strip()
        if current_text and current_text not in self.selected_items:
            self.selected_items.append(current_text)
            self.update_selected_list()

    def remove_selected_item(self, item):
        if item and item.text() in self.selected_items:
            self.selected_items.remove(item.text())
            self.update_selected_list()


    def clear_selection(self):
        self.selected_items = []
        self.update_selected_list()

    def update_selected_list(self):
        self.selected_list.clear()
        for item in self.selected_items:
            self.selected_list.addItem(item)

    def get_selected_items(self):
        return self.selected_items

    def set_selected_items(self, items):
        self.selected_items = items
        self.update_selected_list()


class GroupCountFilterWidget(QWidget):
    """Виджет для фильтрации по количеству групп с операторами >, <, =, >=, <="""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.conditions = []
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        self.setLayout(layout)

        # Заголовок
        title_label = QLabel("Условия фильтрации количества групп:")
        title_label.setStyleSheet("font-weight: bold; color: #2c3e50; margin-bottom: 5px;")
        layout.addWidget(title_label)

        # Layout для ввода нового условия
        condition_layout = QHBoxLayout()

        # ComboBox для выбора оператора
        self.operator_combo = QComboBox()
        self.operator_combo.addItems(["=", ">", "<", ">=", "<="])
        self.operator_combo.setStyleSheet("""
            QComboBox {
                padding: 6px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
                min-width: 60px;
            }
        """)
        condition_layout.addWidget(self.operator_combo)

        # Поле для ввода значения
        self.value_edit = QLineEdit()
        self.value_edit.setValidator(QRegularExpressionValidator(QRegularExpression("[0-9]*")))
        self.value_edit.setPlaceholderText("Введите число")
        self.value_edit.setStyleSheet("""
            QLineEdit {
                padding: 6px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
            }
        """)
        condition_layout.addWidget(self.value_edit)

        # Кнопка добавления условия
        self.add_button = QPushButton("Добавить условие")
        self.add_button.clicked.connect(self.add_condition)
        self.add_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        condition_layout.addWidget(self.add_button)

        layout.addLayout(condition_layout)

        # Список текущих условий
        self.conditions_list = QListWidget()
        self.conditions_list.setStyleSheet("""
            QListWidget {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
                min-height: 80px;
                max-height: 120px;
            }
            QListWidget::item {
                padding: 5px;
                border-bottom: 1px solid #ecf0f1;
            }
            QListWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)
        layout.addWidget(self.conditions_list)

        # Кнопка удаления выбранного условия
        self.remove_button = QPushButton("Удалить выбранное условие")
        self.remove_button.clicked.connect(self.remove_selected_condition)
        self.remove_button.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        layout.addWidget(self.remove_button)

    def add_condition(self):
        """Добавляет новое условие"""
        operator = self.operator_combo.currentText()
        value = self.value_edit.text().strip()

        if not value:
            QMessageBox.warning(self, "Ошибка", "Введите значение для условия")
            return

        try:
            value_int = int(value)
            condition = f"{operator} {value_int}"

            # Проверяем, нет ли уже такого условия
            for existing_condition in self.conditions:
                if existing_condition == condition:
                    QMessageBox.warning(self, "Ошибка", "Такое условие уже добавлено")
                    return

            self.conditions.append(condition)
            self.update_conditions_list()
            self.value_edit.clear()

        except ValueError:
            QMessageBox.warning(self, "Ошибка", "Введите корректное числовое значение")

    def remove_selected_condition(self):
        """Удаляет выбранное условие"""
        current_row = self.conditions_list.currentRow()
        if current_row >= 0 and current_row < len(self.conditions):
            self.conditions.pop(current_row)
            self.update_conditions_list()

    def update_conditions_list(self):
        """Обновляет список условий"""
        self.conditions_list.clear()
        for condition in self.conditions:
            item = QListWidgetItem(condition)
            self.conditions_list.addItem(item)

    def get_conditions(self):
        """Возвращает список условий"""
        return self.conditions

    def set_conditions(self, conditions):
        """Устанавливает условия"""
        self.conditions = conditions
        self.update_conditions_list()


class MultiSelectFilterWidget(QWidget):
    """Виджет для множественного выбора фильтров с чекбоксами"""

    def __init__(self, label, items=None, parent=None):
        super().__init__(parent)
        self.items = items or []
        self.setup_ui(label)

    def setup_ui(self, label):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        self.setLayout(layout)

        # Заголовок
        title_label = QLabel(label)
        title_label.setStyleSheet("font-weight: bold; color: #2c3e50; margin-bottom: 5px;")
        layout.addWidget(title_label)

        # Группа чекбоксов в ScrollArea
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setMinimumHeight(100)
        scroll_area.setMaximumHeight(120)
        scroll_area.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
            }
            QScrollArea > QWidget > QWidget {
                background-color: white;
            }
        """)

        scroll_widget = QWidget()
        self.scroll_layout = QVBoxLayout(scroll_widget)
        self.scroll_layout.setContentsMargins(8, 8, 8, 8)
        self.scroll_layout.setSpacing(2)

        self.checkboxes = {}
        for item in self.items:
            checkbox = QCheckBox(str(item))
            checkbox.setStyleSheet("""
                QCheckBox {
                    spacing: 5px;
                    padding: 3px;
                    background-color: white;
                }
                QCheckBox::indicator {
                    width: 16px;
                    height: 16px;
                }
                QCheckBox::indicator:unchecked {
                    border: 1px solid #bdc3c7;
                    background-color: white;
                    border-radius: 2px;
                }
                QCheckBox::indicator:checked {
                    border: 1px solid #3498db;
                    background-color: #3498db;
                    border-radius: 2px;
                }
            """)
            self.scroll_layout.addWidget(checkbox)
            self.checkboxes[item] = checkbox

        # Добавляем растягивающий элемент чтобы чекбоксы не растягивались
        self.scroll_layout.addStretch(1)

        scroll_area.setWidget(scroll_widget)
        layout.addWidget(scroll_area)

    def get_selected_items(self):
        """Возвращает список выбранных элементов"""
        return [item for item, checkbox in self.checkboxes.items() if checkbox.isChecked()]

    def set_items(self, items):
        """Устанавливает новые элементы"""
        # Очищаем старые чекбоксы
        for i in reversed(range(self.scroll_layout.count())):
            widget = self.scroll_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()

        self.checkboxes = {}
        self.items = items

        # Добавляем новые чекбоксы
        for item in items:
            checkbox = QCheckBox(str(item))
            checkbox.setStyleSheet("""
                QCheckBox {
                    spacing: 5px;
                    padding: 3px;
                    background-color: white;
                }
                QCheckBox::indicator {
                    width: 16px;
                    height: 16px;
                }
                QCheckBox::indicator:unchecked {
                    border: 1px solid #bdc3c7;
                    background-color: white;
                    border-radius: 2px;
                }
                QCheckBox::indicator:checked {
                    border: 1px solid #3498db;
                    background-color: #3498db;
                    border-radius: 2px;
                }
            """)
            self.scroll_layout.insertWidget(self.scroll_layout.count() - 1, checkbox)
            self.checkboxes[item] = checkbox


class GRNTIDialog(QDialog):
    """Диалог для добавления/редактирования кодов ГРНТИ"""

    def __init__(self, parent=None, db=None, expert_id=None, current_grnti=None):
        super().__init__(parent)
        self.db = db
        self.expert_id = expert_id
        self.current_grnti = current_grnti or []
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("Управление кодами ГРНТИ")
        self.setMinimumSize(600, 400)

        layout = QVBoxLayout()

        # Таблица для отображения текущих кодов ГРНТИ
        self.table_label = QLabel("Текущие коды ГРНТИ:")
        self.table_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 12px;")
        layout.addWidget(self.table_label)

        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(3)
        self.table_widget.setHorizontalHeaderLabels(["Код ГРНТИ", "Подрубрика", "Дисциплина"])
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_widget.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_widget.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        layout.addWidget(self.table_widget)

        # Поля для ввода нового кода ГРНТИ
        form_layout = QVBoxLayout()

        grnti_label = QLabel("Добавить новый код ГРНТИ:")
        grnti_label.setStyleSheet("font-weight: bold; color: #2c3e50; margin-top: 10px;")
        form_layout.addWidget(grnti_label)

        # Поле для выбора кода ГРНТИ
        code_layout = QHBoxLayout()
        code_label = QLabel("Код ГРНТИ:")
        code_label.setStyleSheet("font-weight: bold;")
        self.code_combo = QComboBox()
        self.code_combo.setStyleSheet("""
            QComboBox {
                padding: 6px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
            }
            QComboBox::drop-down {
                border: none;
            }
        """)
        self.load_grnti_codes()
        code_layout.addWidget(code_label)
        code_layout.addWidget(self.code_combo)
        form_layout.addLayout(code_layout)

        # Поле для подрубрики (обязательное)
        subrubric_layout = QHBoxLayout()
        subrubric_label = QLabel("Подрубрика*:")
        subrubric_label.setStyleSheet("font-weight: bold;")
        self.subrubric_field = QLineEdit()
        self.subrubric_field.setStyleSheet("""
            QLineEdit {
                padding: 6px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
            }
        """)
        subrubric_layout.addWidget(subrubric_label)
        subrubric_layout.addWidget(self.subrubric_field)
        form_layout.addLayout(subrubric_layout)

        # Поле для дисциплины
        discipline_layout = QHBoxLayout()
        discipline_label = QLabel("Дисциплина:")
        discipline_label.setStyleSheet("font-weight: bold;")
        self.discipline_field = QLineEdit()
        self.discipline_field.setStyleSheet("""
            QLineEdit {
                padding: 6px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
            }
        """)
        discipline_layout.addWidget(discipline_label)
        discipline_layout.addWidget(self.discipline_field)
        form_layout.addLayout(discipline_layout)

        layout.addLayout(form_layout)

        # Кнопки управления - ПЕРЕМЕЩЕНЫ ВНИЗ
        button_layout = QHBoxLayout()

        self.add_button = QPushButton("Добавить")
        self.add_button.clicked.connect(self.add_grnti)
        self.add_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
        """)
        button_layout.addWidget(self.add_button)

        self.remove_button = QPushButton("Удалить")
        self.remove_button.clicked.connect(self.remove_grnti)
        self.remove_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:pressed {
                background-color: #a93226;
            }
        """)
        button_layout.addWidget(self.remove_button)

        # Кнопки OK/Cancel с измененной надписью
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                                   QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

        # Изменяем текст кнопок на русский
        ok_button = buttons.button(QDialogButtonBox.StandardButton.Ok)
        ok_button.setText("ОК")
        cancel_button = buttons.button(QDialogButtonBox.StandardButton.Cancel)
        cancel_button.setText("Отмена")

        # Стилизация кнопок диалога
        for button in buttons.buttons():
            if buttons.buttonRole(button) == QDialogButtonBox.ButtonRole.AcceptRole:
                button.setStyleSheet("""
                    QPushButton {
                        background-color: #27ae60;
                        color: white;
                        border: none;
                        padding: 8px 16px;
                        border-radius: 4px;
                        font-weight: bold;
                        min-width: 80px;
                    }
                    QPushButton:hover {
                        background-color: #219a52;
                    }
                    QPushButton:pressed {
                        background-color: #1e8449;
                    }
                """)
            else:
                button.setStyleSheet("""
                    QPushButton {
                        background-color: #95a5a6;
                        color: white;
                        border: none;
                        padding: 8px 16px;
                        border-radius: 4px;
                        font-weight: bold;
                        min-width: 80px;
                    }
                    QPushButton:hover {
                        background-color: #7f8c8d;
                    }
                    QPushButton:pressed {
                        background-color: #707b7c;
                    }
                """)

        button_layout.addWidget(buttons)
        layout.addLayout(button_layout)

        self.setLayout(layout)
        self.update_table()

    def load_grnti_codes(self):
        """Загружает доступные коды ГРНТИ из базы данных"""
        try:
            if self.db:
                grnti_data = self.db.get_table_data('grnti_classifier')
                for row in grnti_data:
                    code = str(row[0])
                    description = str(row[1]) if len(row) > 1 else ""
                    display_text = f"{code} - {description}" if description else code
                    self.code_combo.addItem(display_text, code)
        except Exception as e:
            print(f"Ошибка загрузки кодов ГРНТИ: {e}")

    def update_table(self):
        """Обновляет таблицу с текущими кодами ГРНТИ"""
        self.table_widget.setRowCount(len(self.current_grnti))
        for row, (code, subrubric, discipline) in enumerate(self.current_grnti):
            # Форматируем код ГРНТИ для отображения
            formatted_code = self.format_grnti_code(str(code))
            code_item = QTableWidgetItem(formatted_code)
            code_item.setFlags(code_item.flags() & ~Qt.ItemFlag.ItemIsEditable)

            # Исправление: отображаем подрубрику и дисциплину, даже если они None
            subrubric_display = str(subrubric) if subrubric is not None else ""
            discipline_display = str(discipline) if discipline is not None else ""

            subrubric_item = QTableWidgetItem(subrubric_display)
            subrubric_item.setFlags(subrubric_item.flags() & ~Qt.ItemFlag.ItemIsEditable)

            discipline_item = QTableWidgetItem(discipline_display)
            discipline_item.setFlags(discipline_item.flags() & ~Qt.ItemFlag.ItemIsEditable)

            self.table_widget.setItem(row, 0, code_item)
            self.table_widget.setItem(row, 1, subrubric_item)
            self.table_widget.setItem(row, 2, discipline_item)

    def format_grnti_code(self, code):
        """Форматирует код ГРНТИ, делая первую часть двузначной"""
        parts = str(code).split('.')
        if len(parts) > 0:
            # Делаем первую часть двузначной
            parts[0] = parts[0].zfill(2)
        return '.'.join(parts)

    def add_grnti(self):
        """Добавляет новый код ГРНТИ"""
        if self.code_combo.currentData() is None:
            QMessageBox.warning(self, "Ошибка", "Выберите код ГРНТИ")
            return

        code = self.code_combo.currentData()
        subrubric = self.subrubric_field.text().strip()
        discipline = self.discipline_field.text().strip()

        # Проверка обязательности подрубрики
        if not subrubric:
            QMessageBox.warning(self, "Ошибка", "Поле 'Подрубрика' обязательно для заполнения")
            self.subrubric_field.setFocus()
            return

        # Проверяем, не добавлена ли уже точно такая же комбинация
        for existing_code, existing_subrubric, existing_discipline in self.current_grnti:
            if (str(existing_code) == str(code) and
                    existing_subrubric == subrubric and
                    existing_discipline == discipline):
                QMessageBox.warning(self, "Ошибка", "Эта комбинация кода ГРНТИ, подрубрики и дисциплины уже добавлена")
                return

        # Если комбинация уникальна - добавляем
        self.current_grnti.append((code, subrubric, discipline))
        self.update_table()

        # Очищаем поля ввода
        self.subrubric_field.clear()
        self.discipline_field.clear()

    def remove_grnti(self):
        """Удаляет выбранный код ГРНТИ"""
        current_row = self.table_widget.currentRow()
        if current_row >= 0 and current_row < len(self.current_grnti):
            self.current_grnti.pop(current_row)
            self.update_table()
        else:
            QMessageBox.warning(self, "Ошибка", "Выберите код ГРНТИ для удаления")

    def get_grnti_data(self):
        """Возвращает данные о кодах ГРНТИ"""
        return self.current_grnti


class ExpertSelectionDialog(QDialog):
    """Диалог для выбора экспертов с фильтрацией по региону и рубрике ГРНТИ"""

    def __init__(self, parent=None, db=None, selected_experts=None, existing_group_members=None):
        super().__init__(parent)
        self.db = db
        self.selected_experts = selected_experts or []  # список (id, name)
        self.existing_group_members = existing_group_members or []  # список ID уже находящихся в группе экспертов
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("Выбор экспертов для группы")
        self.setMinimumSize(800, 600)
        layout = QVBoxLayout()

        # Фильтры
        filter_layout = QHBoxLayout()

        # Регион
        region_label = QLabel("Регион:")
        self.region_combo = QComboBox()
        self.region_combo.addItem("Любой")
        if self.db:
            regions = self.db.get_regions()
            self.region_combo.addItems(regions)
        filter_layout.addWidget(region_label)
        filter_layout.addWidget(self.region_combo)

        # Рубрика ГРНТИ
        rubric_label = QLabel("Рубрика ГРНТИ:")
        self.rubric_combo = QComboBox()
        self.rubric_combo.addItem("Любая")
        if self.db:
            rubrics = self.db.get_unique_rubrics()
            self.rubric_combo.addItems([str(r) for r in sorted(rubrics)])
        filter_layout.addWidget(rubric_label)
        filter_layout.addWidget(self.rubric_combo)

        self.apply_filter_button = QPushButton("Применить фильтр")
        self.apply_filter_button.clicked.connect(self.load_experts)
        filter_layout.addWidget(self.apply_filter_button)

        layout.addLayout(filter_layout)

        # Таблица экспертов
        self.experts_table = QTableWidget()
        self.experts_table.setColumnCount(3)
        self.experts_table.setHorizontalHeaderLabels(["ФИО", "Регион", "Статус"])
        self.experts_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.experts_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.experts_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.experts_table)

        # Кнопка "Добавить выбранного"
        self.add_selected_button = QPushButton("Добавить выбранного эксперта")
        self.add_selected_button.clicked.connect(self.add_selected_expert)
        layout.addWidget(self.add_selected_button)

        # Выбранные эксперты
        selected_label = QLabel("Выбранные эксперты:")
        layout.addWidget(selected_label)

        self.selected_list = QListWidget()
        self.selected_list.setSelectionMode(QListWidget.SelectionMode.SingleSelection)
        self.selected_list.itemDoubleClicked.connect(self.remove_selected_expert)
        layout.addWidget(self.selected_list)

        # OK/Cancel
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                                   QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        ok_button = buttons.button(QDialogButtonBox.StandardButton.Ok)
        ok_button.setText("ОК")
        cancel_button = buttons.button(QDialogButtonBox.StandardButton.Cancel)
        cancel_button.setText("Отмена")
        layout.addWidget(buttons)

        self.setLayout(layout)
        self.load_experts()

    def load_experts(self):
        region = self.region_combo.currentText()
        rubric = self.rubric_combo.currentText()

        region_filter = None if region == "Любой" else region
        rubric_filter = None
        if rubric != "Любая":
            try:
                rubric_filter = int(rubric)
            except ValueError:
                pass

        experts = self.db.get_experts_for_group(region_filter, rubric_filter)
        self.experts_table.setRowCount(len(experts))

        for i, (eid, name, reg, rub) in enumerate(experts):
            self.experts_table.setItem(i, 0, QTableWidgetItem(name or ""))
            self.experts_table.setItem(i, 1, QTableWidgetItem(reg or ""))

            # Проверяем, находится ли эксперт уже в группе
            status = "Уже в группе" if eid in self.existing_group_members else "Доступен"
            status_item = QTableWidgetItem(status)

            # Устанавливаем цвет в зависимости от статуса
            if eid in self.existing_group_members:
                status_item.setBackground(Qt.GlobalColor.lightGray)
                status_item.setForeground(Qt.GlobalColor.darkGray)

            self.experts_table.setItem(i, 2, status_item)
            self.experts_table.item(i, 0).setData(Qt.ItemDataRole.UserRole, eid)
            self.experts_table.item(i, 0).setData(Qt.ItemDataRole.UserRole + 1, rub)

        # Обновляем список выбранных
        self.update_selected_list()

    def add_selected_expert(self):
        row = self.experts_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Ошибка", "Выберите эксперта из таблицы")
            return

        item = self.experts_table.item(row, 0)
        eid = item.data(Qt.ItemDataRole.UserRole)
        name = item.text()
        region = self.experts_table.item(row, 1).text()
        rubric = item.data(Qt.ItemDataRole.UserRole + 1)

        # Проверяем, не находится ли эксперт уже в группе
        if eid in self.existing_group_members:
            QMessageBox.warning(self, "Ошибка", f"Эксперт '{name}' уже находится в группе")
            return

        # Проверяем, не добавлен ли уже эксперт
        if any(e[0] == eid for e in self.selected_experts):
            QMessageBox.warning(self, "Ошибка", f"Эксперт '{name}' уже выбран")
            return

        self.selected_experts.append((eid, name, region, rubric))
        self.update_selected_list()

    def remove_selected_expert(self, item):
        if not item:
            return
        eid = item.data(Qt.ItemDataRole.UserRole)
        if eid is None:
            return
        self.selected_experts = [e for e in self.selected_experts if e[0] != eid]
        self.update_selected_list()

    def update_selected_list(self):
        self.selected_list.clear()
        for eid, name, region, rubric in self.selected_experts:
            display_parts = [name]
            if region:
                display_parts.append(region)
            if rubric:
                display_parts.append(f"Рубрика: {rubric}")
            list_item = QListWidgetItem(" | ".join(display_parts))
            list_item.setData(Qt.ItemDataRole.UserRole, eid)
            self.selected_list.addItem(list_item)

    def get_selected_experts(self):
        return self.selected_experts


class GrntiDetailsDialog(QDialog):
    """Диалог с подробным описанием кодов ГРНТИ"""

    def __init__(self, grnti_details, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Детали ГРНТИ")
        self.setMinimumSize(800, 400)

        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(12)

        if grnti_details:
            # Группируем по базовому коду, чтобы избежать дублирования описаний
            unique_codes = {}
            for detail in grnti_details:
                base_code = detail.get('base_code') or ""
                if not base_code:
                    code_full = detail.get('code_full') or ""
                    parts = [part.strip() for part in code_full.split('.') if part.strip()]
                    base_code = parts[0] if len(parts) > 0 else ""

                if base_code and base_code not in unique_codes:
                    unique_codes[base_code] = {
                        'code': base_code,
                        'description': detail.get('description') or "Описание отсутствует"
                    }

            table = QTableWidget()
            table.setColumnCount(2)
            table.setRowCount(len(unique_codes))
            table.setHorizontalHeaderLabels(["Код ГРНТИ", "Описание"])
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
            table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
            table.setAlternatingRowColors(True)

            for row, (code, data) in enumerate(sorted(unique_codes.items())):
                formatted_code = self.format_grnti_code(code) if code else "—"
                code_item = QTableWidgetItem(formatted_code)
                table.setItem(row, 0, code_item)

                description = data['description']
                description_item = QTableWidgetItem(description)
                description_item.setToolTip(description)
                description_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
                table.setItem(row, 1, description_item)

            table.resizeRowsToContents()
            layout.addWidget(table)
        else:
            empty_label = QLabel("Для выбранного эксперта не указаны коды ГРНТИ.")
            empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            empty_label.setStyleSheet("color: #7f8c8d; font-size: 13px;")
            layout.addWidget(empty_label)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        close_button = buttons.button(QDialogButtonBox.StandardButton.Close)
        close_button.setText("Закрыть")
        close_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 18px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.setLayout(layout)

    def format_grnti_code(self, code):
        """Форматирует код ГРНТИ, делая первую часть двузначной"""
        if not code:
            return ""
        parts = str(code).split('.')
        if len(parts) > 0:
            # Делаем первую часть двузначной
            parts[0] = parts[0].zfill(2)
        return '.'.join(parts)


class UserNameDialog(QDialog):
    """Диалог для ввода ФИО пользователя при экспорте отчета"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Ввод данных пользователя")
        self.setMinimumSize(400, 150)
        
        # Загружаем сохраненное имя пользователя
        settings = QSettings()
        saved_name = settings.value("last_user_name", "")

        layout = QVBoxLayout()

        label = QLabel("Введите ФИО пользователя, создающего отчет:")
        layout.addWidget(label)

        self.name_field = QLineEdit()
        self.name_field.setPlaceholderText("Иванов Иван Иванович")
        if saved_name:
            self.name_field.setText(saved_name)
        layout.addWidget(self.name_field)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        
        # Изменяем текст кнопок на русский
        ok_button = buttons.button(QDialogButtonBox.StandardButton.Ok)
        ok_button.setText("ОК")
        cancel_button = buttons.button(QDialogButtonBox.StandardButton.Cancel)
        cancel_button.setText("Отмена")
        
        layout.addWidget(buttons)

        self.setLayout(layout)

    def get_user_name(self):
        """Возвращает введенное ФИО"""
        name = self.name_field.text().strip()
        # Сохраняем имя пользователя
        if name:
            settings = QSettings()
            settings.setValue("last_user_name", name)
        return name


class GroupMembersDialog(QDialog):
    """Диалог для просмотра состава группы в виде таблицы"""

    def __init__(self, parent=None, db=None, group_id=None, group_name=None, created_at=None):
        super().__init__(parent)
        self.db = db
        self.group_id = group_id
        self.group_name = group_name
        self.created_at = created_at
        self.grnti_details_cache = {}
        self.members_data = []  # Сохраняем данные участников для экспорта
        self.setup_ui()
        self.load_group_members()

    def setup_ui(self):
        self.setWindowTitle(f"Состав группы: {self.group_name}")
        self.setMinimumSize(1000, 600)

        layout = QVBoxLayout()

        # Информация о группе
        info_label = QLabel(f"Группа: {self.group_name}")
        info_label.setStyleSheet("font-weight: bold; font-size: 14px; color: #2c3e50; margin-bottom: 10px;")
        layout.addWidget(info_label)

        # Таблица экспертов
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(6)
        self.table_widget.setHorizontalHeaderLabels([
            "ФИО", "Регион", "Город", "Ключевые слова", "Кол-во групп", "Коды ГРНТИ"
        ])
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_widget.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_widget.setAlternatingRowColors(True)
        self.table_widget.setMouseTracking(True)
        self.table_widget.cellDoubleClicked.connect(self.on_cell_double_clicked)
        self.table_widget.cellEntered.connect(self.on_cell_entered)
        self.table_widget.viewport().installEventFilter(self)
        self.grnti_column_index = 5
        layout.addWidget(self.table_widget)

        # Кнопки экспорта и закрытия
        buttons_layout = QHBoxLayout()

        # Кнопка экспорта в Excel
        self.export_excel_btn = QPushButton("Экспорт в Excel")
        self.export_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #219a52;
            }
        """)
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        buttons_layout.addWidget(self.export_excel_btn)

        # Кнопка экспорта в PDF
        self.export_pdf_btn = QPushButton("Экспорт в PDF")
        self.export_pdf_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        self.export_pdf_btn.clicked.connect(self.export_to_pdf)
        buttons_layout.addWidget(self.export_pdf_btn)

        buttons_layout.addStretch()

        # Кнопка закрытия
        close_button = QPushButton("Закрыть")
        close_button.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        close_button.clicked.connect(self.reject)
        buttons_layout.addWidget(close_button)

        layout.addLayout(buttons_layout)

        self.setLayout(layout)

    def load_group_members(self):
        """Загружает состав группы"""
        try:
            # Получаем экспертов группы с полной информацией
            cursor = self.db.connection.cursor()
            cursor.execute("""
                SELECT 
                    e.id, e.name, e.region, e.city, e.keywords, e.group_count,
                    STRING_AGG(
                        CONCAT_WS('.', 
                            g.codrub::text, 
                            eg.subrubric, 
                            eg.siscipline
                        ), 
                        '; '
                        ORDER BY g.codrub
                    ) AS grnti
                FROM expert e
                JOIN expert_group_link l ON e.id = l.expert_id
                LEFT JOIN expert_grnti eg ON e.id = eg.id
                LEFT JOIN grnti_classifier g ON eg.rubric = g.codrub
                WHERE l.group_id = %s
                GROUP BY e.id, e.name, e.region, e.city, e.keywords, e.group_count
                ORDER BY e.name
            """, (self.group_id,))

            members = cursor.fetchall()
            cursor.close()

            self.table_widget.setRowCount(len(members))
            expert_ids = [row[0] for row in members]
            self.grnti_details_cache = self.db.get_grnti_details_for_experts(expert_ids) if self.db else {}

            # Сохраняем данные участников для экспорта
            self.members_data = []

            for row, (eid, name, region, city, keywords, group_count, grnti) in enumerate(members):
                # Сохраняем данные для экспорта (только ФИО)
                self.members_data.append({
                    'name': name or "",
                    'region': region or "",
                    'city': city or "",
                    'keywords': keywords or "",
                    'group_count': group_count or 0,
                    'grnti': grnti or ""
                })

                name_item = QTableWidgetItem(name or "")
                name_item.setData(Qt.ItemDataRole.UserRole, eid)
                self.table_widget.setItem(row, 0, name_item)
                self.table_widget.setItem(row, 1, QTableWidgetItem(region or ""))
                self.table_widget.setItem(row, 2, QTableWidgetItem(city or ""))
                self.table_widget.setItem(row, 3, QTableWidgetItem(keywords or ""))
                self.table_widget.setItem(row, 4, QTableWidgetItem(str(group_count) if group_count else "0"))

                # Формируем отображение кодов ГРНТИ с возможностью раскрытия подробностей
                grnti_details = self.grnti_details_cache.get(eid, [])
                display_codes = []
                seen_codes = set()
                tooltip_lines = []
                seen_descriptions = {}  # Словарь для отслеживания описаний по базовому коду

                for detail in grnti_details:
                    code_full = detail.get('code_full') or ""
                    parts = [part.strip() for part in code_full.split('.') if part.strip()]

                    base_code = detail.get('base_code') or (parts[0] if len(parts) > 0 else "")
                    rubric_code = detail.get('subrubric') or (parts[1] if len(parts) > 1 else "")
                    subrubric_code = detail.get('discipline') or (parts[2] if len(parts) > 2 else "")

                    formatted_code = self.format_grnti_display(code_full) if code_full else self.format_grnti_display(
                        base_code)
                    if formatted_code and formatted_code not in seen_codes:
                        seen_codes.add(formatted_code)
                        display_codes.append(formatted_code)

                    # Добавляем описание только один раз для каждого базового кода
                    if base_code and base_code not in seen_descriptions:
                        description = detail.get('description') or "Описание отсутствует"
                        seen_descriptions[base_code] = description
                        formatted_base = self.format_grnti_display(base_code)
                        tooltip_line = f"{formatted_base or '—'} — {description}"
                        tooltip_lines.append(tooltip_line)

                if not display_codes and grnti:
                    fallback_codes = [part.strip() for part in str(grnti).split(';') if part.strip()]
                    for code in fallback_codes:
                        formatted_code = self.format_grnti_display(code)
                        if formatted_code and formatted_code not in seen_codes:
                            display_codes.append(formatted_code)

                codes_item = QTableWidgetItem(", ".join(display_codes) if display_codes else "—")
                codes_item.setData(Qt.ItemDataRole.UserRole, grnti_details)
                if tooltip_lines:
                    codes_item.setToolTip("\n\n".join(tooltip_lines))

                if grnti_details:
                    clickable_font = QFont(codes_item.font())
                    clickable_font.setUnderline(True)
                    codes_item.setFont(clickable_font)
                    codes_item.setForeground(QColor("#2980b9"))
                else:
                    codes_item.setForeground(QColor("#34495e"))

                codes_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                self.table_widget.setItem(row, self.grnti_column_index, codes_item)

            self.table_widget.resizeColumnsToContents()

        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось загрузить состав группы: {str(e)}")

    def format_grnti_display(self, value):
        """Форматирует значение, которое может содержать один или несколько кодов ГРНТИ"""
        if not value:
            return ""

        if ';' in value:
            # Несколько кодов
            codes = value.split(';')
            formatted_codes = []
            for code in codes:
                formatted_codes.append(self.format_grnti_code(code.strip()))
            return '; '.join(formatted_codes)
        else:
            return self.format_grnti_code(value)

    def format_grnti_code(self, code):
        """Форматирует код ГРНТИ, делая первую часть двузначной"""
        parts = str(code).split('.')
        if len(parts) > 0:
            # Делаем первую часть двузначной
            parts[0] = parts[0].zfill(2)
        return '.'.join(parts)

    def on_cell_double_clicked(self, row, column):
        if column != self.grnti_column_index:
            return
        item = self.table_widget.item(row, column)
        if not item:
            return

        details = item.data(Qt.ItemDataRole.UserRole) or []
        dialog = GrntiDetailsDialog(details, self)
        dialog.exec()

    def on_cell_entered(self, row, column):
        if column == self.grnti_column_index:
            self.table_widget.viewport().setCursor(Qt.CursorShape.PointingHandCursor)
        else:
            self.table_widget.viewport().setCursor(Qt.CursorShape.ArrowCursor)

    def eventFilter(self, watched, event):
        if watched == self.table_widget.viewport() and event.type() == QEvent.Type.Leave:
            self.table_widget.viewport().setCursor(Qt.CursorShape.ArrowCursor)
        return super().eventFilter(watched, event)

    def export_to_excel(self):
        settings = QSettings()
        """Экспорт отчета в Excel"""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "Ошибка",
                                "Библиотека openpyxl не установлена.\n"
                                "Установите её командой: pip install openpyxl")
            return

        # Запрашиваем ФИО пользователя
        name_dialog = UserNameDialog(self)
        if name_dialog.exec() != QDialog.DialogCode.Accepted:
            return

        user_name = name_dialog.get_user_name()
        if not user_name:
            QMessageBox.warning(self, "Ошибка", "Введите ФИО пользователя")
            return

        # Выбираем файл для сохранения
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить отчет в Excel",
            f"{self.group_name}_отчет.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчет"

            # Стили
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=14)
            title_font = Font(bold=True, size=16)
            normal_font = Font(size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            row = 1

            # Заголовок с номером отчета
            ws.merge_cells(f'A{row}:E{row}')
            title_cell = ws[f'A{row}']
            report_number = settings.value("report_counter", 0, type=int) + 1
            title_cell.value = f"Отчет № {report_number} от {report_date}"
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1

            # Строка "Параметры группы:"
            ws.merge_cells(f'A{row}:E{row}')
            params_cell = ws[f'A{row}']
            params_cell.value = "Параметры группы:"
            params_cell.font = Font(bold=True, size=12)
            params_cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1

            # Информация о группе (сдвигаем остальную информацию ниже)
            ws[f'A{row}'] = "Название:"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = self.group_name
            row += 1

            # Дата создания группы
            if self.created_at:
                created_date_str = self.created_at.strftime('%d.%m.%Y') if hasattr(self.created_at,
                                                                                   'strftime') else str(self.created_at)
            else:
                created_date_str = "Не указана"

            ws[f'A{row}'] = "Дата создания группы:"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = created_date_str
            row += 1

            # Дата формирования отчета
            report_date = datetime.now().strftime('%d.%m.%Y %H:%M')
            ws[f'A{row}'] = "Дата формирования отчета:"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = report_date
            row += 2

            # Заголовки таблицы
            headers = ["№", "ФИО участника", "Регион", "Город", "Коды ГРНТИ"]
            header_row = row
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            row += 1

            # Данные участников
            for idx, member in enumerate(self.members_data, 1):
                # Номер
                ws.cell(row=row, column=1, value=idx).border = border
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='top')

                # ФИО
                ws.cell(row=row, column=2, value=member.get('name', '') or '').border = border
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                # Регион
                ws.cell(row=row, column=3, value=member.get('region', '') or '—').border = border
                ws.cell(row=row, column=3).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                # Город
                ws.cell(row=row, column=4, value=member.get('city', '') or '—').border = border
                ws.cell(row=row, column=4).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                # Коды ГРНТИ (каждый код с новой строки)
                grnti_codes = member.get('grnti', '') or ''
                if grnti_codes:
                    codes_list = [code.strip() for code in str(grnti_codes).split(';') if code.strip()]
                    grnti_display = '\n'.join(codes_list)  # Каждый код с новой строки
                else:
                    grnti_display = "—"

                ws.cell(row=row, column=5, value=grnti_display).border = border
                ws.cell(row=row, column=5).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                row += 1

            # Настройка ширины столбцов
            ws.column_dimensions['A'].width = 8  # №
            ws.column_dimensions['B'].width = 35  # ФИО
            ws.column_dimensions['C'].width = 25  # Регион
            ws.column_dimensions['D'].width = 25  # Город
            ws.column_dimensions['E'].width = 30  # Коды ГРНТИ

            # Высота строк заголовка
            ws.row_dimensions[header_row].height = 25

            row += 2

            # ФИО пользователя, создавшего отчет
            ws[f'A{row}'] = "Отчет создал:"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = user_name
            row += 2

            # Поле для подписи
            ws[f'A{row}'] = "Подпись:"
            ws[f'A{row}'].font = Font(bold=True)
            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'].border = Border(
                bottom=Side(style='medium')
            )
            ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='bottom')

            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Отчет успешно сохранен в файл:\n{file_path}")

        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось сохранить отчет:\n{str(e)}")

    def export_to_pdf(self):
        """Экспорт отчета в PDF"""
        if not REPORTLAB_AVAILABLE:
            QMessageBox.warning(self, "Ошибка",
                                "Библиотека reportlab не установлена.\n"
                                "Установите её командой: pip install reportlab")
            return

        # Запрашиваем ФИО пользователя
        name_dialog = UserNameDialog(self)
        if name_dialog.exec() != QDialog.DialogCode.Accepted:
            return

        user_name = name_dialog.get_user_name()
        if not user_name:
            QMessageBox.warning(self, "Ошибка", "Введите ФИО пользователя")
            return

        # Выбираем файл для сохранения
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить отчет в PDF",
            f"{self.group_name}_отчет.pdf",
            "PDF Files (*.pdf)"
        )

        if not file_path:
            return

        # Получаем и увеличиваем сквозной номер отчета (только после выбора файла)
        settings = QSettings()
        report_number = settings.value("report_counter", 0, type=int)
        report_number += 1

        try:
            # Регистрируем шрифты с поддержкой кириллицы
            # Пытаемся использовать системные шрифты
            cyrillic_font_name = 'CyrillicFont'
            cyrillic_font_bold_name = 'CyrillicFontBold'

            # Пробуем найти системные шрифты с поддержкой кириллицы
            font_paths = [
                # Windows
                'C:/Windows/Fonts/arial.ttf',
                'C:/Windows/Fonts/arialbd.ttf',
                'C:/Windows/Fonts/times.ttf',
                'C:/Windows/Fonts/timesbd.ttf',
                # macOS - более полный список
                '/System/Library/Fonts/Supplemental/Arial.ttf',
                '/System/Library/Fonts/Supplemental/Arial Bold.ttf',
                '/Library/Fonts/Arial.ttf',
                '/Library/Fonts/Arial Bold.ttf',
                '/System/Library/Fonts/Helvetica.ttc',
                '/System/Library/Fonts/HelveticaNeue.ttc',
                # Linux
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
                '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
                '/usr/share/fonts/truetype/ttf-dejavu/DejaVuSans.ttf',
                '/usr/share/fonts/truetype/ttf-dejavu/DejaVuSans-Bold.ttf',
            ]

            regular_font_found = False
            bold_font_found = False

            for font_path in font_paths:
                if os.path.exists(font_path):
                    # Пропускаем .ttc файлы, так как TTFont их не поддерживает напрямую
                    if font_path.endswith('.ttc'):
                        continue
                    try:
                        if 'bold' in font_path.lower() or 'Bold' in font_path or 'bd' in font_path.lower():
                            if not bold_font_found:
                                pdfmetrics.registerFont(TTFont(cyrillic_font_bold_name, font_path))
                                bold_font_found = True
                        else:
                            if not regular_font_found:
                                pdfmetrics.registerFont(TTFont(cyrillic_font_name, font_path))
                                regular_font_found = True
                    except Exception as e:
                        # Пропускаем файлы, которые не удалось загрузить
                        continue

                    if regular_font_found and bold_font_found:
                        break

            # Если не нашли системные шрифты, пробуем использовать встроенные шрифты reportlab
            # или используем стандартные с fallback на Helvetica
            if not regular_font_found:
                # Пробуем использовать встроенные шрифты с поддержкой Unicode
                try:
                    # Используем стандартные шрифты, но с правильной кодировкой
                    cyrillic_font_name = 'Helvetica'
                except:
                    cyrillic_font_name = 'Helvetica'

            if not bold_font_found:
                try:
                    cyrillic_font_bold_name = 'Helvetica-Bold'
                except:
                    cyrillic_font_bold_name = 'Helvetica-Bold'

            doc = SimpleDocTemplate(file_path, pagesize=A4,
                                    leftMargin=10 * mm, rightMargin=10 * mm,
                                    topMargin=15 * mm, bottomMargin=15 * mm)
            story = []

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=cyrillic_font_bold_name,
                fontSize=18,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=30,
                alignment=1  # center
            )

            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontName=cyrillic_font_bold_name,
                fontSize=14,
                textColor=colors.HexColor('#34495e'),
                spaceAfter=12,
                spaceBefore=12
            )

            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName=cyrillic_font_name,
                fontSize=11
            )

            # Определяем дату отчета (дата создания группы)
            if self.created_at:
                report_date_str = self.created_at.strftime('%d.%m.%Y') if hasattr(self.created_at,
                                                                                   'strftime') else str(self.created_at)
            else:
                report_date_str = datetime.now().strftime('%d.%m.%Y')

            # Заголовок: Отчет <номер> от <дата отчета>
            title = Paragraph(f"Отчет № {report_number} от {report_date_str}", title_style)
            story.append(title)
            story.append(Spacer(1, 15))

            # Строка "Параметры группы:"
            params_heading = Paragraph("<b>Параметры группы:</b>", heading_style)
            story.append(params_heading)
            story.append(Spacer(1, 10))

            # Информация о группе
            story.append(Paragraph(f"<b>Название:</b> {self.group_name}", normal_style))
            story.append(Spacer(1, 5))
            
            # Дата создания группы
            if self.created_at:
                created_date_str = self.created_at.strftime('%d.%m.%Y') if hasattr(self.created_at,
                                                                                   'strftime') else str(self.created_at)
            else:
                created_date_str = "Не указана"
            
            story.append(Paragraph(f"<b>Дата создания:</b> {created_date_str}", normal_style))
            story.append(Spacer(1, 5))
            
            # Количество участников
            participants_count = len(self.members_data)
            story.append(Paragraph(f"<b>Кол-во участников:</b> {participants_count}", normal_style))
            story.append(Spacer(1, 20))

            # Заголовок таблицы
            story.append(Paragraph("<b>Состав группы:</b>", heading_style))

            # Таблица участников
            # Создаем стили для ячеек таблицы
            cell_style = ParagraphStyle(
                'TableCell',
                parent=normal_style,
                fontName=cyrillic_font_name,
                fontSize=8,
                leading=10,
                alignment=0  # LEFT
            )

            cell_header_style = ParagraphStyle(
                'TableHeader',
                parent=normal_style,
                fontName=cyrillic_font_bold_name,
                fontSize=9,
                alignment=0  # LEFT
            )

            table_data = []
            # Заголовки таблицы
            header_row = [
                Paragraph("№", cell_header_style),
                Paragraph("ФИО участника", cell_header_style),
                Paragraph("Регион", cell_header_style),
                Paragraph("Город", cell_header_style),
                Paragraph("Коды ГРНТИ", cell_header_style)
            ]
            table_data.append(header_row)

            for idx, member in enumerate(self.members_data, 1):
                # Форматируем коды ГРНТИ для отображения в столбец
                grnti_codes = member.get('grnti', '') or ''
                if grnti_codes:
                    # Разбиваем коды и отображаем каждый с новой строки
                    codes_list = [code.strip() for code in str(grnti_codes).split(';') if code.strip()]
                    # Каждый код на новой строке
                    grnti_display = '<br/>'.join(codes_list)
                else:
                    grnti_display = "—"

                # Используем Paragraph для переноса текста
                name_text = member.get('name', '') or '—'
                region_text = member.get('region', '') or '—'
                city_text = member.get('city', '') or '—'

                # Заменяем длинные пробелы на неразрывные пробелы для лучшего переноса
                name_text = name_text.replace('  ', ' ')
                region_text = region_text.replace('  ', ' ')
                city_text = city_text.replace('  ', ' ')

                table_data.append([
                    Paragraph(str(idx), cell_style),  # Номер
                    Paragraph(name_text, cell_style),  # ФИО
                    Paragraph(region_text, cell_style),  # Регион
                    Paragraph(city_text, cell_style),  # Город
                    Paragraph(grnti_display, cell_style)  # Коды ГРНТИ (вертикально)
                ])

            # Оптимизируем ширины колонок для A4 (ширина ~210mm минус отступы по 10mm = 190mm доступно)
            # №: 10mm, ФИО: 45mm, Регион: 30mm, Город: 30mm, ГРНТИ: 75mm = 190mm
            table = Table(table_data, colWidths=[10 * mm, 45 * mm, 30 * mm, 30 * mm, 75 * mm])
            table.hAlign = 'LEFT'  # Выравнивание таблицы по левому краю страницы
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8E8E8')),  # Светлый фон вместо синего
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Черный текст вместо белого
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Только номер по центру
                ('ALIGN', (1, 0), (-1, -1), 'LEFT'),  # Остальные колонки по левому краю
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), cyrillic_font_bold_name),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                ('TOPPADDING', (0, 0), (-1, 0), 5),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E6F3FF')),  # Светло-голубой фон вместо желтого
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), cyrillic_font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E6F3FF')]),  # Светло-голубой вместо желтого
            ]))

            story.append(table)
            story.append(Spacer(1, 30))

            # ФИО пользователя
            story.append(Paragraph(f"<b>Отчет создал:</b> {user_name}", normal_style))
            story.append(Spacer(1, 20))

            # Поле для подписи
            story.append(Paragraph("<b>Подпись:</b>", normal_style))
            story.append(Spacer(1, 40))
            story.append(Paragraph("_" * 50, normal_style))
            story.append(Spacer(1, 20))
            
            # Поле для даты подписи (оставляем место для заполнения вручную)
            story.append(Paragraph("<b>Дата подписи:</b> _________________", normal_style))

            doc.build(story)
            # Сохраняем счетчик только после успешного сохранения файла
            settings.setValue("report_counter", report_number)
            QMessageBox.information(self, "Успех", f"Отчет успешно сохранен в файл:\n{file_path}")

        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось сохранить отчет:\n{str(e)}")


class GroupEditDialog(QDialog):
    def __init__(self, parent=None, db=None, is_edit=False, group_data=None, group_id=None):
        super().__init__(parent)
        self.db = db
        self.is_edit = is_edit
        self.group_id = group_id
        self.selected_experts = []  # [(id, name, region, rubric), ...] - БЕЗ ДУБЛИКАТОВ
        self.existing_group_members = []  # ID уже находящихся в группе экспертов
        self.original_experts = []  # Сохраняем исходный состав для сравнения
        self.setup_ui()
        if is_edit and group_data:
            self.fill_data(group_data)

    def setup_ui(self):
        self.setWindowTitle("Редактирование группы" if self.is_edit else "Создание группы")
        self.setMinimumSize(600, 500)
        layout = QVBoxLayout()

        # Название
        layout.addWidget(QLabel("Название группы*"))
        self.name_field = QLineEdit()
        layout.addWidget(self.name_field)

        # Дата создания — только для чтения
        layout.addWidget(QLabel("Дата создания"))
        self.date_field = QLineEdit()
        self.date_field.setReadOnly(True)
        if not self.is_edit:
            self.date_field.setText(datetime.now().strftime('%d.%m.%Y'))
        layout.addWidget(self.date_field)

        # Автополя (только для просмотра)
        layout.addWidget(QLabel("Регионы участников (авто):"))
        self.auto_regions_field = QLineEdit()
        self.auto_regions_field.setReadOnly(True)
        layout.addWidget(self.auto_regions_field)

        layout.addWidget(QLabel("Рубрики ГРНТИ участников (авто):"))
        self.auto_grnti_field = QLineEdit()
        self.auto_grnti_field.setReadOnly(True)
        layout.addWidget(self.auto_grnti_field)

        # Кнопка выбора экспертов
        self.select_experts_button = QPushButton("Выбрать экспертов")
        self.select_experts_button.clicked.connect(self.open_expert_selection)
        self.select_experts_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        layout.addWidget(self.select_experts_button)

        # Надпись с количеством выбранных экспертов
        self.selected_experts_label = QLabel("Выбрано экспертов: 0")
        self.selected_experts_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        layout.addWidget(self.selected_experts_label)

        # Список выбранных экспертов с возможностью удаления
        layout.addWidget(QLabel("Текущие участники группы:"))
        self.experts_list = QListWidget()
        self.experts_list.setSelectionMode(QListWidget.SelectionMode.SingleSelection)
        self.experts_list.setMinimumHeight(150)
        layout.addWidget(self.experts_list)

        # Кнопка удаления выбранного эксперта из группы
        self.remove_expert_button = QPushButton("Удалить выбранного эксперта из группы")
        self.remove_expert_button.clicked.connect(self.remove_selected_expert)
        self.remove_expert_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
        """)
        layout.addWidget(self.remove_expert_button)

        # OK/Cancel
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                                   QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.validate_and_accept)
        buttons.rejected.connect(self.reject)
        ok_button = buttons.button(QDialogButtonBox.StandardButton.Ok)
        ok_button.setText("ОК")
        cancel_button = buttons.button(QDialogButtonBox.StandardButton.Cancel)
        cancel_button.setText("Отмена")
        layout.addWidget(buttons)

        self.setLayout(layout)

        # Обновляем состояние кнопки удаления
        self.update_remove_button_state()

    def open_expert_selection(self):
        # Получаем ID уже находящихся в группе экспертов (БЕЗ ДУБЛИКАТОВ)
        existing_member_ids = list(set([expert[0] for expert in self.selected_experts]))

        dialog = ExpertSelectionDialog(self, self.db, self.selected_experts.copy(), existing_member_ids)
        if dialog.exec():
            new_selected_experts = dialog.get_selected_experts()

            # УБИРАЕМ ДУБЛИКАТЫ по expert_id
            seen_ids = set()
            self.selected_experts = []
            for expert in new_selected_experts:
                if expert[0] not in seen_ids:
                    self.selected_experts.append(expert)
                    seen_ids.add(expert[0])

            self.update_participant_count()
            self.update_auto_fields()
            self.update_experts_list()
            self.update_remove_button_state()

    def update_participant_count(self):
        count = len(self.selected_experts)
        self.selected_experts_label.setText(f"Выбрано экспертов: {count}")

    def update_auto_fields(self):
        """Автоматически заполняет регионы и рубрики на основе выбранных экспертов"""
        if not self.selected_experts or not self.db:
            self.auto_regions_field.setText("")
            self.auto_grnti_field.setText("")
            return

        expert_ids = [item[0] for item in self.selected_experts]
        regions = set()
        grnti_rubrics = set()

        cursor = self.db.connection.cursor()
        try:
            # Получаем регионы
            cursor.execute("SELECT DISTINCT region FROM expert WHERE id = ANY(%s)", (expert_ids,))
            regions.update(row[0] for row in cursor.fetchall() if row[0])

            # Получаем ВСЕ рубрики ГРНТИ для выбранных экспертов
            cursor.execute("""
                SELECT DISTINCT rubric 
                FROM expert_grnti 
                WHERE id = ANY(%s) AND rubric IS NOT NULL
                ORDER BY rubric
            """, (expert_ids,))

            rubrics_data = cursor.fetchall()
            for row in rubrics_data:
                if row[0] is not None:
                    grnti_rubrics.add(str(row[0]))

            print(f"Автополя: регионы={list(regions)}, рубрики={list(grnti_rubrics)}")

        except Exception as e:
            print(f"Ошибка получения автополей: {e}")
        finally:
            cursor.close()

        self.auto_regions_field.setText(", ".join(sorted(regions)) if regions else "—")
        self.auto_grnti_field.setText(", ".join(sorted(grnti_rubrics)) if grnti_rubrics else "—")

    def update_experts_list(self):
        """Обновляет список экспертов в интерфейсе (БЕЗ ДУБЛИКАТОВ)"""
        self.experts_list.clear()
        for expert_id, name, region, rubric in self.selected_experts:
            parts = [name]
            if region:
                parts.append(region)
            if rubric:
                parts.append(f"Рубрика: {rubric}")
            item = QListWidgetItem(" | ".join(parts))
            item.setData(Qt.ItemDataRole.UserRole, expert_id)
            self.experts_list.addItem(item)

    def update_remove_button_state(self):
        """Обновляет состояние кнопки удаления в зависимости от наличия выбранных экспертов"""
        has_experts = len(self.selected_experts) > 0
        self.remove_expert_button.setEnabled(has_experts)

    def remove_selected_expert(self):
        """Удаляет выбранного эксперта из группы"""
        current_row = self.experts_list.currentRow()
        if current_row >= 0 and current_row < len(self.selected_experts):
            # Подтверждение удаления
            expert_name = self.selected_experts[current_row][1]
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Подтверждение удаления")
            msg_box.setText(f"Вы уверены, что хотите удалить эксперта '{expert_name}' из группы?\n\n"
                           f"Счетчик групп у эксперта будет уменьшен на 1.")
            msg_box.setIcon(QMessageBox.Icon.Question)
            
            # Создаем кнопки с русским текстом
            yes_button = msg_box.addButton("Да", QMessageBox.ButtonRole.YesRole)
            no_button = msg_box.addButton("Нет", QMessageBox.ButtonRole.NoRole)
            msg_box.setDefaultButton(no_button)
            
            msg_box.exec()
            
            if msg_box.clickedButton() == yes_button:
                self.selected_experts.pop(current_row)
                self.update_participant_count()
                self.update_auto_fields()
                self.update_experts_list()
                self.update_remove_button_state()
        else:
            QMessageBox.warning(self, "Ошибка", "Выберите эксперта для удаления")

    def fill_data(self, group_data):
        # group_data: (id, name, participant_count, created_at, regions, rubric_code)
        if len(group_data) < 6:
            QMessageBox.warning(self, "Ошибка", "Неполные данные группы")
            return

        self.name_field.setText(group_data[1] or "")
        # Отображаем дату создания из БД (не текущую!)
        self.date_field.setText(DateValidator.format_date_for_display(str(group_data[3])) or "")

        # Загружаем участников и обновляем UI (БЕЗ ДУБЛИКАТОВ)
        if self.group_id:
            all_members = self.db.get_group_members_with_details(self.group_id)
            # Убираем дубликаты по expert_id
            seen_ids = set()
            self.selected_experts = []
            for member in all_members:
                if member[0] not in seen_ids:
                    self.selected_experts.append(member)
                    seen_ids.add(member[0])

            # Сохраняем исходный состав для сравнения
            self.original_experts = self.selected_experts.copy()

            self.update_participant_count()
            self.update_auto_fields()
            self.update_experts_list()
            self.update_remove_button_state()

    def validate_and_accept(self):
        if not self.name_field.text().strip():
            QMessageBox.warning(self, "Ошибка", "Укажите название группы")
            return
        # УДАЛЕНА ПРОВЕРКА НА НАЛИЧИЕ ЭКСПЕРТОВ - ТЕПЕРЬ МОЖНО СОЗДАВАТЬ ПУСТУЮ ГРУППУ
        self.accept()

    def get_data(self):
        regions = sorted({e[2] for e in self.selected_experts if e[2]})

        # Получаем ВСЕ рубрики ГРНТИ для выбранных экспертов из БД
        grnti_rubrics = set()
        if self.selected_experts and self.db:
            cursor = self.db.connection.cursor()
            try:
                expert_ids = [e[0] for e in self.selected_experts]
                if expert_ids:
                    # Получаем ВСЕ уникальные рубрики для выбранных экспертов
                    cursor.execute("""
                        SELECT DISTINCT eg.rubric 
                        FROM expert_grnti eg
                        WHERE eg.id = ANY(%s) AND eg.rubric IS NOT NULL
                        ORDER BY eg.rubric
                    """, (expert_ids,))

                    rubrics_data = cursor.fetchall()
                    for row in rubrics_data:
                        rubric = row[0]
                        if rubric is not None:
                            grnti_rubrics.add(str(rubric))

                    print(f"Найдено рубрик для экспертов {expert_ids}: {list(grnti_rubrics)}")

            except Exception as e:
                print(f"Ошибка получения рубрик: {e}")
            finally:
                cursor.close()

        # Уже без дубликатов благодаря обработке в open_expert_selection
        unique_experts = [(e[0], e[1]) for e in self.selected_experts]

        result_rubrics = ", ".join(sorted(grnti_rubrics)) if grnti_rubrics else None

        print(f"Итоговые рубрики для группы: {result_rubrics}")
        print(f"Количество экспертов: {len(unique_experts)}")

        return {
            'name': self.name_field.text().strip(),
            'participant_count': len(unique_experts),
            'regions': ", ".join(regions) if regions else None,
            'rubric_code': result_rubrics,
            'created_at': datetime.now().strftime('%Y-%m-%d'),
            'experts': unique_experts,
            'original_experts': self.original_experts  # Добавляем для сравнения изменений
        }


class AddToGroupDialog(QDialog):
    """Диалог для добавления экспертов в группу - КОМБИНИРОВАННАЯ ВЕРСИЯ"""

    def __init__(self, parent=None, db=None, expert_ids=None, expert_names=None):
        super().__init__(parent)
        self.db = db
        self.expert_ids = expert_ids or []
        self.expert_names = expert_names or []
        self.all_groups = []  # Все группы для фильтрации
        self.selected_group_id = None  # Храним выбранный ID группы
        self.filtered_groups = []  # Отфильтрованные группы
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("Добавить экспертов в группу")
        self.setMinimumSize(750, 650)
        self.resize(750, 650)
        
        # Устанавливаем светлый фон для всего диалога и всех виджетов
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QWidget {
                background-color: #f8f9fa;
            }
            QScrollArea {
                background-color: #f8f9fa;
            }
            QLabel {
                background-color: transparent;
            }
        """)

        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # Блок 1: Информация о выбранных экспертах
        experts_group = QGroupBox("Выбранные эксперты")
        experts_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                background-color: #f8f9fa;
            }
        """)
        experts_layout = QVBoxLayout()
        experts_layout.setSpacing(8)
        experts_layout.setContentsMargins(12, 15, 12, 12)

        # Счетчик экспертов
        info_label = QLabel(f"Количество: {len(self.expert_ids)}")
        info_label.setStyleSheet("font-weight: bold; font-size: 12px; color: #34495e; margin-bottom: 5px;")
        experts_layout.addWidget(info_label)

        # Полный список экспертов с прокруткой
        if self.expert_names:
            experts_scroll = QScrollArea()
            experts_scroll.setWidgetResizable(True)
            experts_scroll.setMaximumHeight(120)
            experts_scroll.setStyleSheet("""
                QScrollArea {
                    border: 1px solid #dee2e6;
                    border-radius: 6px;
                    background-color: white;
                }
            """)

            experts_widget = QWidget()
            experts_list_layout = QVBoxLayout(experts_widget)
            experts_list_layout.setSpacing(3)
            experts_list_layout.setContentsMargins(8, 8, 8, 8)

            for name in self.expert_names:
                expert_label = QLabel(f"• {name}")
                expert_label.setStyleSheet("margin: 2px; padding: 4px; font-size: 11px; color: #555;")
                experts_list_layout.addWidget(expert_label)

            experts_scroll.setWidget(experts_widget)
            experts_layout.addWidget(experts_scroll)
        else:
            no_experts_label = QLabel("Эксперты не выбраны")
            no_experts_label.setStyleSheet("color: #95a5a6; font-style: italic; padding: 10px;")
            experts_layout.addWidget(no_experts_label)

        experts_group.setLayout(experts_layout)
        layout.addWidget(experts_group)

        # Блок 2: Выбор действия
        action_group = QGroupBox("Действие")
        action_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 2px solid #27ae60;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                background-color: #f8f9fa;
            }
        """)
        action_layout = QVBoxLayout()
        action_layout.setSpacing(10)
        action_layout.setContentsMargins(12, 15, 12, 12)

        self.action_combo = QComboBox()

        # Загружаем группы и определяем, есть ли существующие
        has_existing_groups = self.load_groups()

        # Добавляем опции в зависимости от наличия групп
        if has_existing_groups:
            self.action_combo.addItem("Добавить в существующую группу", "existing")
            self.action_combo.addItem("Создать новую группу", "new")
        else:
            self.action_combo.addItem("Создать новую группу", "new")
            # Показываем информационное сообщение
            info_label = QLabel("Существующие группы отсутствуют. Будет создана новая группа.")
            info_label.setStyleSheet(
                "color: #856404; background-color: #fff3cd; padding: 10px; border-radius: 6px; font-size: 11px; border: 1px solid #ffc107;")
            action_layout.addWidget(info_label)

        self.action_combo.currentIndexChanged.connect(self.on_action_changed)
        self.action_combo.setStyleSheet("""
            QComboBox {
                padding: 10px;
                border: 2px solid #27ae60;
                border-radius: 6px;
                background-color: white;
                color: #2c3e50;
                font-size: 13px;
                font-weight: 500;
            }
            QComboBox:hover {
                border-color: #219a52;
            }
            QComboBox::drop-down {
                border: none;
                width: 30px;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                color: #2c3e50;
                font-size: 12px;
                border: 2px solid #27ae60;
                border-radius: 6px;
                padding: 5px;
                selection-background-color: #27ae60;
                selection-color: white;
            }
        """)
        action_layout.addWidget(self.action_combo)
        action_group.setLayout(action_layout)
        layout.addWidget(action_group)

        # Блок 3: Выбор существующей группы
        self.existing_group_widget = QGroupBox("Выбор существующей группы")
        self.existing_group_widget.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 2px solid #9b59b6;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                background-color: #f8f9fa;
            }
        """)
        existing_layout = QVBoxLayout()
        existing_layout.setSpacing(12)
        existing_layout.setContentsMargins(12, 15, 12, 12)

        # Поиск группы
        search_label = QLabel("Поиск по названию, дате или количеству участников:")
        search_label.setStyleSheet("font-weight: 600; font-size: 12px; color: #34495e; margin-bottom: 5px;")
        existing_layout.addWidget(search_label)

        # Поле поиска с выпадающим списком результатов
        self.search_field = QLineEdit()
        self.search_field.setPlaceholderText("Начните вводить название группы, дату или количество участников...")
        self.search_field.textChanged.connect(self.on_search_text_changed)
        self.search_field.setStyleSheet("""
            QLineEdit {
                padding: 10px;
                border: 2px solid #9b59b6;
                border-radius: 6px;
                background-color: white;
                color: #2c3e50;
                font-size: 12px;
            }
            QLineEdit:focus {
                border-color: #8e44ad;
                background-color: #fff;
            }
        """)
        existing_layout.addWidget(self.search_field)

        # Выпадающий список для отображения найденных групп
        self.results_list = QListWidget()
        self.results_list.setVisible(False)  # Сначала скрыт
        self.results_list.setMaximumHeight(150)
        self.results_list.itemClicked.connect(self.on_result_selected)
        self.results_list.setStyleSheet("""
            QListWidget {
                border: 2px solid #9b59b6;
                border-radius: 6px;
                background-color: white;
                color: #2c3e50;
                font-size: 12px;
            }
            QListWidget::item {
                padding: 10px;
                border-bottom: 1px solid #ecf0f1;
            }
            QListWidget::item:selected {
                background-color: #9b59b6;
                color: white;
                border-radius: 4px;
            }
            QListWidget::item:hover {
                background-color: #e8d5f0;
            }
        """)
        existing_layout.addWidget(self.results_list)

        # Разделитель между поиском и списком
        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.HLine)
        separator.setFrameShadow(QFrame.Shadow.Sunken)
        separator.setStyleSheet("background-color: #dee2e6; margin: 10px 0;")
        existing_layout.addWidget(separator)

        # ВЫБОР ГРУППЫ ИЗ СПИСКА
        select_label = QLabel("Или выберите группу из списка:")
        select_label.setStyleSheet("font-weight: 600; font-size: 12px; color: #34495e; margin-top: 5px; margin-bottom: 5px;")
        existing_layout.addWidget(select_label)

        self.group_combo = QComboBox()
        self.group_combo.currentIndexChanged.connect(self.on_group_combo_changed)
        self.group_combo.setStyleSheet("""
            QComboBox {
                padding: 10px;
                border: 2px solid #9b59b6;
                border-radius: 6px;
                background-color: white;
                color: #2c3e50;
                font-size: 12px;
            }
            QComboBox:hover {
                border-color: #8e44ad;
            }
            QComboBox::drop-down {
                border: none;
                width: 30px;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                color: #2c3e50;
                font-size: 12px;
                border: 2px solid #9b59b6;
                border-radius: 6px;
                padding: 5px;
                selection-background-color: #9b59b6;
                selection-color: white;
                max-height: 200px;
            }
        """)

        # Заполняем комбобокс группами
        self.update_group_combo()
        existing_layout.addWidget(self.group_combo)

        self.existing_group_widget.setLayout(existing_layout)
        layout.addWidget(self.existing_group_widget)

        # Блок 4: Создание новой группы
        self.new_group_widget = QGroupBox("Создание новой группы")
        self.new_group_widget.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 2px solid #e74c3c;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                background-color: #f8f9fa;
            }
        """)
        new_layout = QVBoxLayout()
        new_layout.setSpacing(10)
        new_layout.setContentsMargins(12, 15, 12, 12)

        new_label = QLabel("Введите название новой группы:")
        new_label.setStyleSheet("font-weight: 600; font-size: 12px; color: #34495e; margin-bottom: 5px;")
        new_layout.addWidget(new_label)

        self.new_group_name = QLineEdit()
        self.new_group_name.setPlaceholderText("Например: Группа экспертов по информатике")
        self.new_group_name.setStyleSheet("""
            QLineEdit {
                padding: 10px;
                border: 2px solid #e74c3c;
                border-radius: 6px;
                background-color: white;
                color: #2c3e50;
                font-size: 12px;
            }
            QLineEdit:focus {
                border-color: #c0392b;
                background-color: #fff;
            }
        """)
        new_layout.addWidget(self.new_group_name)
        self.new_group_widget.setLayout(new_layout)
        layout.addWidget(self.new_group_widget)

        # Изначально скрываем оба виджета и покажем нужный после проверки
        if has_existing_groups:
            self.existing_group_widget.setVisible(True)
            self.new_group_widget.setVisible(False)
        else:
            self.existing_group_widget.setVisible(False)
            self.new_group_widget.setVisible(True)
            self.action_combo.setCurrentIndex(0)  # Автоматически выбираем создание новой группы

        # Добавляем растягивающий элемент перед кнопками
        layout.addStretch(1)

        # Кнопки
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                                   QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.validate_and_accept)
        buttons.rejected.connect(self.reject)

        # Изменяем текст кнопок на русский
        ok_button = buttons.button(QDialogButtonBox.StandardButton.Ok)
        ok_button.setText("ОК")
        cancel_button = buttons.button(QDialogButtonBox.StandardButton.Cancel)
        cancel_button.setText("Отмена")

        # Стилизация кнопок
        for button in buttons.buttons():
            if buttons.buttonRole(button) == QDialogButtonBox.ButtonRole.AcceptRole:
                button.setStyleSheet("""
                    QPushButton {
                        background-color: #27ae60;
                        color: white;
                        border: none;
                        padding: 12px 24px;
                        border-radius: 6px;
                        font-weight: bold;
                        min-width: 100px;
                        font-size: 13px;
                    }
                    QPushButton:hover {
                        background-color: #219a52;
                    }
                    QPushButton:pressed {
                        background-color: #1e8449;
                    }
                """)
            else:
                button.setStyleSheet("""
                    QPushButton {
                        background-color: #95a5a6;
                        color: white;
                        border: none;
                        padding: 12px 24px;
                        border-radius: 6px;
                        font-weight: bold;
                        min-width: 100px;
                        font-size: 13px;
                    }
                    QPushButton:hover {
                        background-color: #7f8c8d;
                    }
                    QPushButton:pressed {
                        background-color: #707b7c;
                    }
                """)

        layout.addWidget(buttons)
        self.setLayout(layout)

        # Устанавливаем фокус на поле поиска, если виджет видим
        if has_existing_groups:
            QTimer.singleShot(100, self.search_field.setFocus)

    def update_group_combo(self):
        """Обновляет комбобокс со списком групп"""
        self.group_combo.clear()
        self.group_combo.addItem("Выберите группу...", None)

        for group in self.all_groups:
            group_id, name, participant_count, created_at, regions, rubric_code = group
            date_display = self.format_date_for_display(created_at)
            display_text = f"{name} (Участников: {participant_count}, Дата: {date_display})"
            self.group_combo.addItem(display_text, group_id)

    def on_group_combo_changed(self, index):
        """Обработчик изменения выбора в комбобоксе"""
        if index > 0:  # Пропускаем первый элемент "Выберите группу..."
            group_id = self.group_combo.currentData()
            if group_id:
                self.selected_group_id = group_id

                # Находим группу в all_groups для отображения
                for group in self.all_groups:
                    if group[0] == group_id:
                        group_name = group[1]

                        # Также заполняем поле поиска для согласованности
                        self.search_field.blockSignals(True)
                        self.search_field.setText(group_name)
                        self.search_field.blockSignals(False)
                        break

    def load_groups(self):
        """Загружает список групп и возвращает True если группы есть"""
        groups = []
        try:
            groups = self.db.get_table_data('expert_group') or []
        except Exception as e:
            print(f"Ошибка загрузки групп: {e}")
            groups = []

        # Сохраняем все группы для фильтрации
        self.all_groups = groups
        self.filtered_groups = groups.copy()  # Изначально показываем все группы

        has_groups = len(groups) > 0
        return has_groups

    def on_search_text_changed(self, search_text):
        """Обработчик изменения текста поиска"""
        if not search_text.strip():
            # Если поле поиска пустое, скрываем выпадающий список
            self.results_list.setVisible(False)
            self.filtered_groups = self.all_groups.copy()
            return

        # Фильтруем группы
        filtered_groups = self.filter_groups(search_text)
        self.filtered_groups = filtered_groups

        # Обновляем выпадающий список результатов
        self.update_results_list(filtered_groups)

        # Показываем или скрываем список в зависимости от результатов
        has_results = len(filtered_groups) > 0
        self.results_list.setVisible(has_results)

    def filter_groups(self, search_text):
        """Фильтрует группы по названию или дате"""
        if not search_text.strip():
            return self.all_groups

        filtered_groups = []
        search_lower = search_text.lower()

        for group in self.all_groups:
            group_id, name, participant_count, created_at, regions, rubric_code = group

            # Поиск по названию (регистронезависимый)
            name_match = search_lower in (name or "").lower()

            # Поиск по дате
            date_match = False
            if created_at:
                date_match = self.check_date_match(created_at, search_text)

            # Поиск по регионам (если есть)
            region_match = False
            if regions:
                region_match = search_lower in (regions or "").lower()

            # Поиск по количеству участников
            count_match = False
            if participant_count is not None:
                count_match = search_text in str(participant_count)

            if name_match or date_match or region_match or count_match:
                filtered_groups.append(group)

        return filtered_groups

    def check_date_match(self, created_at, search_text):
        """Проверяет совпадение даты с поисковым запросом"""
        try:
            # Форматируем дату для поиска в различных форматах
            if isinstance(created_at, (date, datetime)):
                date_formats = [
                    created_at.strftime('%d.%m.%Y'),  # 11.11.2025
                    created_at.strftime('%d.%m.%y'),  # 11.11.25
                    created_at.strftime('%m.%Y'),  # 11.2025
                    created_at.strftime('%m.%y'),  # 11.25
                    created_at.strftime('%Y'),  # 2025
                    created_at.strftime('%d.%m'),  # 11.11 (без года)
                    created_at.strftime('%Y-%m-%d'),  # 2025-11-11
                    created_at.strftime('%Y/%m/%d'),  # 2025/11/11
                    created_at.strftime('%d-%m-%Y'),  # 11-11-2025
                    created_at.strftime('%d/%m/%Y'),  # 11/11/2025
                ]
            else:
                date_obj = DateValidator.parse_date(str(created_at))
                if date_obj:
                    date_formats = [
                        date_obj.strftime('%d.%m.%Y'),
                        date_obj.strftime('%d.%m.%y'),
                        date_obj.strftime('%m.%Y'),
                        date_obj.strftime('%m.%y'),
                        date_obj.strftime('%Y'),
                        date_obj.strftime('%d.%m'),
                        date_obj.strftime('%Y-%m-%d'),
                        date_obj.strftime('%Y/%m/%d'),
                        date_obj.strftime('%d-%m-%Y'),
                        date_obj.strftime('%d/%m/%Y'),
                    ]
                else:
                    date_formats = [str(created_at)]

            # Проверяем все форматы даты
            for date_str in date_formats:
                if search_text in date_str:
                    return True

        except Exception as e:
            print(f"Ошибка обработки даты: {e}")

        return False

    def update_results_list(self, groups):
        """Обновляет список результатов поиска"""
        self.results_list.clear()

        if not groups:
            item = QListWidgetItem("Группы не найдены")
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEnabled)  # Делаем недоступным для выбора
            self.results_list.addItem(item)
            return

        for group in groups:
            group_id, name, participant_count, created_at, regions, rubric_code = group

            # Форматируем дату для отображения
            date_display = self.format_date_for_display(created_at)

            display_text = f"{name} (Участников: {participant_count}, Дата: {date_display})"
            item = QListWidgetItem(display_text)
            item.setData(Qt.ItemDataRole.UserRole, group_id)  # Сохраняем ID группы
            self.results_list.addItem(item)

    def format_date_for_display(self, created_at):
        """Форматирует дату для отображения"""
        if created_at:
            if hasattr(created_at, 'strftime'):
                return created_at.strftime('%d.%m.%Y')
            else:
                try:
                    date_obj = DateValidator.parse_date(str(created_at))
                    return date_obj.strftime('%d.%m.%Y') if date_obj else str(created_at)
                except:
                    return str(created_at)
        else:
            return "—"

    def on_result_selected(self, item):
        """Обработчик выбора результата из списка"""
        if not item.flags() & Qt.ItemFlag.ItemIsEnabled:
            return  # Пропускаем недоступные элементы

        group_id = item.data(Qt.ItemDataRole.UserRole)
        if group_id:
            self.selected_group_id = group_id

            # Находим группу в all_groups
            for group in self.all_groups:
                if group[0] == group_id:
                    group_name = group[1]
                    # Заполняем поле поиска названием группы
                    self.search_field.blockSignals(True)
                    self.search_field.setText(group_name)
                    self.search_field.blockSignals(False)

                    # Устанавливаем соответствующий элемент в комбобоксе
                    for i in range(self.group_combo.count()):
                        if self.group_combo.itemData(i) == group_id:
                            self.group_combo.setCurrentIndex(i)
                            break

                    # Скрываем выпадающий список
                    self.results_list.setVisible(False)
                    break

    def on_action_changed(self):
        """Обработчик изменения выбора действия"""
        action = self.action_combo.currentData()
        if action == "existing":
            self.existing_group_widget.setVisible(True)
            self.new_group_widget.setVisible(False)
            # Устанавливаем фокус на поле поиска при переключении
            QTimer.singleShot(100, self.search_field.setFocus)
        else:
            self.existing_group_widget.setVisible(False)
            self.new_group_widget.setVisible(True)

    def validate_and_accept(self):
        """Проверяет валидность данных перед принятием"""
        action = self.action_combo.currentData()

        if action == "existing":
            if self.selected_group_id is None:
                QMessageBox.warning(self, "Ошибка", "Выберите группу для добавления экспертов")
                return
        else:
            group_name = self.new_group_name.text().strip()
            if not group_name:
                QMessageBox.warning(self, "Ошибка", "Введите название для новой группы")
                return

            # Проверяем, что название не слишком длинное
            if len(group_name) > 100:
                QMessageBox.warning(self, "Ошибка", "Название группы не должно превышать 100 символов")
                return

        self.accept()

    def get_selected_action(self):
        """Возвращает выбранное действие и данные"""
        action = self.action_combo.currentData()
        if action == "existing":
            # Находим название выбранной группы
            group_name = ""
            for group in self.all_groups:
                if group[0] == self.selected_group_id:
                    group_name = group[1]
                    break

            return {
                'action': 'existing',
                'group_id': self.selected_group_id,
                'group_name': group_name
            }
        else:
            return {
                'action': 'new',
                'group_name': self.new_group_name.text().strip()
            }

    def keyPressEvent(self, event):
        """Обработка нажатий клавиш для улучшения UX"""
        if event.key() == Qt.Key.Key_Escape and self.results_list.isVisible():
            # Скрываем выпадающий список при нажатии Escape
            self.results_list.setVisible(False)
            event.accept()
        elif event.key() == Qt.Key.Key_Down and self.search_field.hasFocus() and not self.results_list.isVisible():
            # Показываем выпадающий список при нажатии стрелки вниз в поле поиска
            if self.filtered_groups:
                self.results_list.setVisible(True)
            event.accept()
        else:
            super().keyPressEvent(event)

    def focusOutEvent(self, event):
        """Обработчик потери фокуса - скрываем выпадающий список"""
        # Проверяем, был ли клик вне выпадающего списка
        from PyQt6.QtGui import QCursor
        if (not self.results_list.geometry().contains(self.mapFromGlobal(QCursor.pos())) and
                not self.search_field.geometry().contains(self.mapFromGlobal(QCursor.pos()))):
            self.results_list.setVisible(False)
        super().focusOutEvent(event)


class FilterDialog(QDialog):
    """Диалог для фильтрации данных в общей таблице"""

    def __init__(self, parent=None, db=None, current_filters=None):
        super().__init__(parent)
        self.db = db
        self.current_filters = current_filters or {}
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("Фильтрация экспертов")
        self.setMinimumSize(800, 700)  # Увеличенный размер для новых элементов

        # Создаем основную прокручиваемую область
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        main_layout = QVBoxLayout()
        main_layout.addWidget(scroll_area)
        self.setLayout(main_layout)

        # Создаем виджет для содержимого
        content_widget = QWidget()
        scroll_area.setWidget(content_widget)

        layout = QVBoxLayout(content_widget)
        layout.setSpacing(12)
        layout.setContentsMargins(15, 15, 15, 15)

        # Информация о фильтрации
        info_label = QLabel("Заполните одно или несколько полей для фильтрации")
        info_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 14px; margin-bottom: 15px;")
        layout.addWidget(info_label)

        # Основные фильтры
        basic_group = QGroupBox("Основные фильтры")
        basic_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px 0 8px;
            }
        """)
        basic_layout = QVBoxLayout()
        basic_layout.setSpacing(10)
        basic_layout.setContentsMargins(12, 12, 12, 12)

        # ФИО эксперта - комбобокс с автодополнением и множественным выбором
        name_label = QLabel("ФИО эксперта:")
        name_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 12px;")
        basic_layout.addWidget(name_label)

        # Получаем список всех ФИО экспертов
        expert_names = []
        if self.db:
            try:
                experts_data = self.db.get_table_data('expert')
                for row in experts_data:
                    if len(row) > 1 and row[1]:  # name field
                        expert_names.append(str(row[1]))
            except Exception as e:
                print(f"Ошибка загрузки ФИО экспертов: {e}")

        self.name_widget = MultiSelectComboBox(sorted(list(set(expert_names))), self)
        basic_layout.addWidget(self.name_widget)

        # Регион - комбобокс с автодополнением и множественным выбором
        region_label = QLabel("Регион:")
        region_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 12px;")
        basic_layout.addWidget(region_label)

        regions = self.db.get_regions() if self.db else []
        self.region_widget = MultiSelectComboBox(regions, self)
        basic_layout.addWidget(self.region_widget)

        # Город - комбобокс с автодополнением и множественным выбором
        city_label = QLabel("Город:")
        city_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 12px;")
        basic_layout.addWidget(city_label)

        # Получаем все города
        all_cities = []
        if self.db:
            try:
                cities_data = self.db.get_all_cities_with_regions()
                for city, region in cities_data:
                    if city:  # Проверяем, что город не пустой
                        all_cities.append(city)
            except Exception as e:
                print(f"Ошибка загрузки городов: {e}")

        self.city_widget = MultiSelectComboBox(sorted(list(set(all_cities))), self)
        basic_layout.addWidget(self.city_widget)

        # Ключевые слова
        keywords_label = QLabel("Ключевые слова:")
        keywords_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 12px;")
        basic_layout.addWidget(keywords_label)
        self.keywords_field = QLineEdit()
        self.keywords_field.setPlaceholderText("Введите ключевые слова через запятую")
        self.keywords_field.setStyleSheet("""
            QLineEdit {
                padding: 10px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
                font-size: 12px;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
        """)
        basic_layout.addWidget(self.keywords_field)

        # Количество групп - новый виджет с операторами
        group_count_label = QLabel("Количество групп (можно задать условия >, <, =, >=, <=):")
        group_count_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 12px;")
        basic_layout.addWidget(group_count_label)

        self.group_count_widget = GroupCountFilterWidget(self)
        basic_layout.addWidget(self.group_count_widget)

        basic_group.setLayout(basic_layout)
        layout.addWidget(basic_group)

        # Фильтры по специализации
        specialization_group = QGroupBox("Фильтры по специализации")
        specialization_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px 0 8px;
            }
        """)
        specialization_layout = QVBoxLayout()
        specialization_layout.setSpacing(10)
        specialization_layout.setContentsMargins(12, 12, 12, 12)

        # Загружаем данные для фильтров
        grnti_codes = []
        subrubrics = set()
        disciplines = set()

        if self.db:
            try:
                # Загружаем коды ГРНТИ
                grnti_data = self.db.get_table_data('grnti_classifier')
                for row in grnti_data:
                    code = str(row[0])
                    description = str(row[1]) if len(row) > 1 else ""
                    display_text = f"{code} - {description}" if description else code
                    grnti_codes.append(display_text)

                # Загружаем уникальные подрубрики и дисциплины
                expert_grnti_data = self.db.get_table_data('expert_grnti')
                for row in expert_grnti_data:
                    if len(row) > 2 and row[2] is not None:  # subrubric
                        subrubrics.add(str(row[2]))
                    if len(row) > 3 and row[3] is not None:  # discipline
                        disciplines.add(str(row[3]))
            except Exception as e:
                print(f"Ошибка загрузки данных для фильтров: {e}")

        # Код ГРНТИ - множественный выбор
        grnti_label = QLabel("Коды ГРНТИ:")
        grnti_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 12px;")
        specialization_layout.addWidget(grnti_label)

        self.grnti_widget = MultiSelectComboBox(grnti_codes, self)
        specialization_layout.addWidget(self.grnti_widget)

        # Подрубрика - множественный выбор
        subrubric_label = QLabel("Подрубрики:")
        subrubric_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 12px;")
        specialization_layout.addWidget(subrubric_label)

        self.subrubric_widget = MultiSelectComboBox(sorted(list(subrubrics)), self)
        specialization_layout.addWidget(self.subrubric_widget)

        # Дисциплина - множественный выбор
        discipline_label = QLabel("Дисциплины:")
        discipline_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 12px;")
        specialization_layout.addWidget(discipline_label)

        self.discipline_widget = MultiSelectComboBox(sorted(list(disciplines)), self)
        specialization_layout.addWidget(self.discipline_widget)

        specialization_group.setLayout(specialization_layout)
        layout.addWidget(specialization_group)

        # Заполняем текущие значения фильтров
        self.fill_current_filters()

        # Кнопки - ВНЕ прокручиваемой области
        button_widget = QWidget()
        button_layout = QVBoxLayout(button_widget)
        button_layout.setContentsMargins(0, 10, 0, 0)

        action_button_layout = QHBoxLayout()

        self.apply_button = QPushButton("Применить фильтр")
        self.apply_button.clicked.connect(self.apply_filters)
        self.apply_button.setStyleSheet("""
            QPushButton { 
                background-color: #3498db; 
                color: white; 
                border: none;
                padding: 12px 24px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
        """)
        action_button_layout.addWidget(self.apply_button)

        self.reset_button = QPushButton("Сбросить фильтр")
        self.reset_button.clicked.connect(self.reset_filters)
        self.reset_button.setStyleSheet("""
            QPushButton { 
                background-color: #95a5a6; 
                color: white; 
                border: none;
                padding: 12px 24px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
            QPushButton:pressed {
                background-color: #707b7c;
            }
        """)
        action_button_layout.addWidget(self.reset_button)

        button_layout.addLayout(action_button_layout)

        # Кнопки OK/Cancel
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                                   QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

        # Изменяем текст кнопок на русский
        ok_button = buttons.button(QDialogButtonBox.StandardButton.Ok)
        ok_button.setText("ОК")
        cancel_button = buttons.button(QDialogButtonBox.StandardButton.Cancel)
        cancel_button.setText("Отмена")

        # Стилизация кнопок диалога
        for button in buttons.buttons():
            if buttons.buttonRole(button) == QDialogButtonBox.ButtonRole.AcceptRole:
                button.setStyleSheet("""
                    QPushButton {
                        background-color: #27ae60;
                        color: white;
                        border: none;
                        padding: 10px 20px;
                        border-radius: 6px;
                        font-weight: bold;
                        min-width: 90px;
                        font-size: 13px;
                    }
                    QPushButton:hover {
                        background-color: #219a52;
                    }
                    QPushButton:pressed {
                        background-color: #1e8449;
                    }
                """)
            else:
                button.setStyleSheet("""
                    QPushButton {
                        background-color: #95a5a6;
                        color: white;
                        border: none;
                        padding: 10px 20px;
                        border-radius: 6px;
                        font-weight: bold;
                        min-width: 90px;
                        font-size: 13px;
                    }
                    QPushButton:hover {
                        background-color: #7f8c8d;
                    }
                    QPushButton:pressed {
                        background-color: #707b7c;
                    }
                """)

        button_layout.addWidget(buttons)
        layout.addWidget(button_widget)

    def fill_current_filters(self):
        """Заполняет поля текущими значениями фильтров"""
        if self.current_filters:
            # ФИО эксперта
            expert_names = self.current_filters.get('expert_names', [])
            if expert_names:
                self.name_widget.set_selected_items(expert_names)

            # Регионы
            regions = self.current_filters.get('regions', [])
            if regions:
                self.region_widget.set_selected_items(regions)

            # Города
            cities = self.current_filters.get('cities', [])
            if cities:
                self.city_widget.set_selected_items(cities)

            self.keywords_field.setText(self.current_filters.get('keywords', ''))

            # Количество групп - новые условия
            group_conditions = self.current_filters.get('group_conditions', [])
            if group_conditions:
                self.group_count_widget.set_conditions(group_conditions)

            # Коды ГРНТИ
            grnti_codes = self.current_filters.get('grnti_codes', [])
            if grnti_codes:
                # Преобразуем коды в строки для отображения
                grnti_codes_str = [str(code) for code in grnti_codes]
                self.grnti_widget.set_selected_items(grnti_codes_str)

            # Подрубрики
            subrubrics = self.current_filters.get('subrubrics', [])
            if subrubrics:
                # Преобразуем подрубрики в строки для отображения
                subrubrics_str = [str(sr) for sr in subrubrics]
                self.subrubric_widget.set_selected_items(subrubrics_str)

            # Дисциплины
            disciplines = self.current_filters.get('disciplines', [])
            if disciplines:
                # Преобразуем дисциплины в строки для отображения
                disciplines_str = [str(d) for d in disciplines]
                self.discipline_widget.set_selected_items(disciplines_str)

    def apply_filters(self):
        """Применяет фильтры и закрывает диалог"""
        filters = self.get_filters()
        self.accept()

    def reset_filters(self):
        """Сбрасывает все фильтры"""
        self.name_widget.clear_selection()
        self.region_widget.clear_selection()
        self.city_widget.clear_selection()
        self.keywords_field.clear()
        self.group_count_widget.set_conditions([])
        self.grnti_widget.clear_selection()
        self.subrubric_widget.clear_selection()
        self.discipline_widget.clear_selection()

    def get_filters(self):
        """Возвращает словарь с установленными фильтрами"""
        filters = {}

        # ФИО эксперта - множественный выбор
        expert_names_selected = self.name_widget.get_selected_items()
        if expert_names_selected:
            filters['expert_names'] = expert_names_selected

        # Регионы - множественный выбор
        regions_selected = self.region_widget.get_selected_items()
        if regions_selected:
            filters['regions'] = regions_selected

        # Города - множественный выбор
        cities_selected = self.city_widget.get_selected_items()
        if cities_selected:
            filters['cities'] = cities_selected

        # Ключевые слова
        keywords = self.keywords_field.text().strip()
        if keywords:
            filters['keywords'] = keywords

        # Количество групп - новые условия
        group_conditions = self.group_count_widget.get_conditions()
        if group_conditions:
            filters['group_conditions'] = group_conditions

        # Коды ГРНТИ - из множественного выбора
        grnti_selected = self.grnti_widget.get_selected_items()
        if grnti_selected:
            codes = []
            for code_text in grnti_selected:
                # Извлекаем числовую часть до дефиса
                code_str = str(code_text).split(' - ')[0].strip()

                # Преобразуем в числовой формат
                try:
                    # Убираем точки для правильного преобразования
                    code_clean = code_str.replace('.', '')
                    code_int = int(code_clean)
                    codes.append(code_int)
                except (ValueError, TypeError) as e:
                    print(f"Ошибка преобразования кода ГРНТИ '{code_str}': {e}")
                    continue

            if codes:
                filters['grnti_codes'] = codes

        # Подрубрики - из множественного выбора
        subrubric_selected = self.subrubric_widget.get_selected_items()
        if subrubric_selected:
            subrubrics = []
            for subrubric in subrubric_selected:
                try:
                    subrubric_int = int(subrubric)
                    subrubrics.append(subrubric_int)
                except (ValueError, TypeError) as e:
                    print(f"Ошибка преобразования подрубрики '{subrubric}': {e}")
                    continue
            if subrubrics:
                filters['subrubrics'] = subrubrics

        # Дисциплины - из множественного выбора
        discipline_selected = self.discipline_widget.get_selected_items()
        if discipline_selected:
            disciplines = []
            for discipline in discipline_selected:
                try:
                    discipline_int = int(discipline)
                    disciplines.append(discipline_int)
                except (ValueError, TypeError) as e:
                    print(f"Ошибка преобразования дисциплины '{discipline}': {e}")
                    continue
            if disciplines:
                filters['disciplines'] = disciplines

        return filters


class DatabaseManager:
    def __init__(self):
        try:
            self.connection = psycopg2.connect(**DB_CONFIG)
            self.connection.autocommit = True
            print("Успешное подключение к базе данных!")
        except Exception as e:
            print(f"Ошибка подключения: {e}")
            raise

    def update_expert_group_count(self, expert_id):
        """Обновляет счетчик групп для конкретного эксперта"""
        cursor = self.connection.cursor()
        try:
            cursor.execute("""
                UPDATE expert 
                SET group_count = (
                    SELECT COUNT(DISTINCT group_id) 
                    FROM expert_group_link 
                    WHERE expert_id = %s
                )
                WHERE id = %s
            """, (expert_id, expert_id))
            self.connection.commit()
            print(f"Обновлен счетчик групп для эксперта {expert_id}")
        except Exception as e:
            self.connection.rollback()
            print(f"Ошибка обновления счетчика групп для эксперта {expert_id}: {e}")
            raise e
        finally:
            cursor.close()

    def update_experts_group_counts(self, expert_ids):
        """Обновляет счетчики групп для нескольких экспертов"""
        if not expert_ids:
            return

        cursor = self.connection.cursor()
        try:
            for expert_id in expert_ids:
                cursor.execute("""
                    UPDATE expert 
                    SET group_count = (
                        SELECT COUNT(DISTINCT group_id) 
                        FROM expert_group_link 
                        WHERE expert_id = %s
                    )
                    WHERE id = %s
                """, (expert_id, expert_id))
                print(f"Обновлен счетчик групп для эксперта {expert_id}")

            self.connection.commit()
            print(f"Обновлены счетчики групп для {len(expert_ids)} экспертов")
        except Exception as e:
            self.connection.rollback()
            print(f"Ошибка обновления счетчиков групп для экспертов {expert_ids}: {e}")
            raise e
        finally:
            cursor.close()

    def decrease_experts_group_count(self, expert_ids):
        """Уменьшает счетчик групп на 1 для указанных экспертов - НОВЫЙ МЕТОД"""
        if not expert_ids:
            return

        cursor = self.connection.cursor()
        try:
            for expert_id in expert_ids:
                cursor.execute("""
                    UPDATE expert 
                    SET group_count = GREATEST(0, group_count - 1)
                    WHERE id = %s
                """, (expert_id,))
                print(f"Уменьшен счетчик групп для эксперта {expert_id}")

            self.connection.commit()
            print(f"Уменьшены счетчики групп для {len(expert_ids)} экспертов")
        except Exception as e:
            self.connection.rollback()
            print(f"Ошибка уменьшения счетчиков групп для экспертов {expert_ids}: {e}")
            raise e
        finally:
            cursor.close()

    def add_experts_to_group(self, group_id, expert_ids):
        """Добавляет экспертов в группу и обновляет их счетчики"""
        cursor = self.connection.cursor()
        try:
            # Получаем текущих участников группы
            cursor.execute("SELECT expert_id FROM expert_group_link WHERE group_id = %s", (group_id,))
            existing_members = [row[0] for row in cursor.fetchall()]

            added_experts = []
            for expert_id in expert_ids:
                # Проверяем, нет ли уже такого эксперта в группе
                if expert_id in existing_members:
                    print(f"Эксперт {expert_id} уже находится в группе {group_id}")
                    continue

                cursor.execute(
                    "INSERT INTO expert_group_link (expert_id, group_id) VALUES (%s, %s)",
                    (expert_id, group_id)
                )
                added_experts.append(expert_id)
                print(f"Добавлен эксперт {expert_id} в группу {group_id}")

            # Обновляем количество участников в группе
            if added_experts:
                cursor.execute(
                    "UPDATE expert_group SET participant_count = participant_count + %s WHERE id = %s",
                    (len(added_experts), group_id)
                )

                # ОБНОВЛЯЕМ СЧЕТЧИКИ ГРУПП ДЛЯ ВСЕХ ДОБАВЛЕННЫХ ЭКСПЕРТОВ
                self.update_experts_group_counts(added_experts)

                # ОБНОВЛЯЕМ поля регионов и рубрик после добавления экспертов
                self.update_group_fields_after_change(group_id)

            self.connection.commit()
            print(f"Добавлено {len(added_experts)} экспертов в группу {group_id}")
            return len(added_experts)
        except Exception as e:
            self.connection.rollback()
            print(f"Ошибка добавления экспертов в группу: {e}")
            raise e
        finally:
            cursor.close()

    def remove_expert_from_group(self, expert_id, group_id):
        """Удаляет эксперта из группы и обновляет его счетчик"""
        cursor = self.connection.cursor()
        try:
            cursor.execute(
                "DELETE FROM expert_group_link WHERE expert_id = %s AND group_id = %s",
                (expert_id, group_id)
            )

            if cursor.rowcount > 0:
                # ОБНОВЛЯЕМ СЧЕТЧИК ГРУПП ДЛЯ ЭКСПЕРТА
                self.update_expert_group_count(expert_id)

                # Обновляем количество участников в группе
                cursor.execute(
                    "UPDATE expert_group SET participant_count = participant_count - 1 WHERE id = %s",
                    (group_id,)
                )

                # ОБНОВЛЯЕМ поля регионов и рубрик после удаления эксперта
                self.update_group_fields_after_change(group_id)

                self.connection.commit()
                print(f"Удален эксперт {expert_id} из группы {group_id}")
                return True
            return False

        except Exception as e:
            self.connection.rollback()
            print(f"Ошибка удаления эксперта из группы: {e}")
            raise e
        finally:
            cursor.close()

    def remove_experts_from_group(self, group_id, expert_ids):
        """Удаляет нескольких экспертов из группы и обновляет их счетчики"""
        if not expert_ids:
            return 0

        cursor = self.connection.cursor()
        try:
            removed_count = 0
            for expert_id in expert_ids:
                cursor.execute(
                    "DELETE FROM expert_group_link WHERE expert_id = %s AND group_id = %s",
                    (expert_id, group_id)
                )
                if cursor.rowcount > 0:
                    removed_count += 1

            if removed_count > 0:
                # ОБНОВЛЯЕМ СЧЕТЧИКИ ГРУПП ДЛЯ ВСЕХ УДАЛЕННЫХ ЭКСПЕРТОВ
                self.update_experts_group_counts(expert_ids)

                # Обновляем количество участников в группе
                cursor.execute(
                    "UPDATE expert_group SET participant_count = participant_count - %s WHERE id = %s",
                    (removed_count, group_id)
                )

                # ОБНОВЛЯЕМ поля регионов и рубрик после удаления экспертов
                self.update_group_fields_after_change(group_id)

            self.connection.commit()
            print(f"Удалено {removed_count} экспертов из группы {group_id}")
            return removed_count
        except Exception as e:
            self.connection.rollback()
            print(f"Ошибка удаления экспертов из группы: {e}")
            raise e
        finally:
            cursor.close()

    def delete_group(self, group_id):
        """Удаляет группу - ПРОСТАЯ И РАБОЧАЯ ВЕРСИЯ"""
        cursor = self.connection.cursor()
        try:
            # Получаем всех экспертов группы перед удалением
            cursor.execute("SELECT expert_id FROM expert_group_link WHERE group_id = %s", (group_id,))
            expert_ids = [row[0] for row in cursor.fetchall()]

            print(f"Удаление группы {group_id} с {len(expert_ids)} экспертами")

            # Удаляем связи экспертов с группой
            cursor.execute("DELETE FROM expert_group_link WHERE group_id = %s", (group_id,))
            link_deleted_count = cursor.rowcount

            # Удаляем саму группу
            cursor.execute("DELETE FROM expert_group WHERE id = %s", (group_id,))
            group_deleted = cursor.rowcount > 0

            # КРИТИЧЕСКОЕ ИСПРАВЛЕНИЕ: обновляем счетчики ПРАВИЛЬНЫМ способом
            if group_deleted and expert_ids:
                # Используем ТОЧНЫЙ подсчет текущего количества групп для каждого эксперта
                for expert_id in expert_ids:
                    cursor.execute("""
                        UPDATE expert 
                        SET group_count = (
                            SELECT COUNT(*) 
                            FROM expert_group_link 
                            WHERE expert_id = %s
                        )
                        WHERE id = %s
                    """, (expert_id, expert_id))
                    print(f"Обновлен счетчик групп для эксперта {expert_id}")

            self.connection.commit()
            print(f"Группа {group_id} удалена, удалено {link_deleted_count} связей с экспертами")
            print(f"Обновлены счетчики групп для {len(expert_ids)} экспертов")
            return group_deleted
        except Exception as e:
            self.connection.rollback()
            print(f"Ошибка удаления группы {group_id}: {e}")
            raise e
        finally:
            cursor.close()

    def create_group_with_experts(self, group_name, expert_ids):
        """Создает новую группу и добавляет в нее экспертов"""
        cursor = self.connection.cursor()
        try:
            print(f"Создание группы '{group_name}' с экспертами: {expert_ids}")

            # Сначала получаем регионы и рубрики для заполнения полей
            regions = set()
            rubrics = set()

            if expert_ids:
                # Получаем регионы экспертов
                cursor.execute("SELECT DISTINCT region FROM expert WHERE id = ANY(%s) AND region IS NOT NULL",
                               (expert_ids,))
                regions.update(row[0] for row in cursor.fetchall() if row[0])

                # Получаем рубрики экспертов
                cursor.execute("""
                    SELECT DISTINCT rubric 
                    FROM expert_grnti 
                    WHERE id = ANY(%s) AND rubric IS NOT NULL
                """, (expert_ids,))
                rubrics.update(str(row[0]) for row in cursor.fetchall() if row[0] is not None)

            regions_str = ", ".join(sorted(regions)) if regions else None
            rubrics_str = ", ".join(sorted(rubrics)) if rubrics else None

            print(f"Регионы группы: {regions_str}")
            print(f"Рубрики группы: {rubrics_str}")

            # Создаем группу с заполненными полями
            actual_expert_count = len(expert_ids)
            cursor.execute(
                "INSERT INTO expert_group (name, participant_count, created_at, regions, rubric_code) VALUES (%s, %s, %s, %s, %s) RETURNING id",
                (group_name, actual_expert_count, datetime.now().strftime('%Y-%m-%d'), regions_str, rubrics_str)
            )
            group_id = cursor.fetchone()[0]
            print(f"Создана группа с ID: {group_id}")

            # Добавляем экспертов в группу (без дубликатов)
            added_experts = set()
            for expert_id in expert_ids:
                if expert_id not in added_experts:
                    cursor.execute(
                        "INSERT INTO expert_group_link (expert_id, group_id) VALUES (%s, %s)",
                        (expert_id, group_id)
                    )
                    added_experts.add(expert_id)
                    print(f"Добавлен эксперт {expert_id} в группу {group_id}")

            # ОБНОВЛЯЕМ СЧЕТЧИКИ ГРУПП ДЛЯ ВСЕХ ДОБАВЛЕННЫХ ЭКСПЕРТОВ
            if added_experts:
                self.update_experts_group_counts(list(added_experts))

            # Двойная проверка - обновляем количество участников на основе реально добавленных
            actual_added_count = len(added_experts)
            if actual_added_count != actual_expert_count:
                cursor.execute(
                    "UPDATE expert_group SET participant_count = %s WHERE id = %s",
                    (actual_added_count, group_id)
                )
                print(f"Обновлено количество участников: {actual_added_count}")

            self.connection.commit()
            print(f"Успешно создана группа '{group_name}' с {actual_added_count} экспертами")
            return group_id
        except Exception as e:
            self.connection.rollback()
            print(f"Ошибка создания группы: {e}")
            raise e
        finally:
            cursor.close()

    def update_group_fields_after_change(self, group_id):
        """Обновляет поля регионов и рубрик после изменений в составе группы"""
        cursor = self.connection.cursor()
        try:
            # Получаем всех экспертов группы
            cursor.execute("SELECT expert_id FROM expert_group_link WHERE group_id = %s", (group_id,))
            expert_ids = [row[0] for row in cursor.fetchall()]

            if not expert_ids:
                return

            # Получаем регионы экспертов
            cursor.execute("SELECT DISTINCT region FROM expert WHERE id = ANY(%s) AND region IS NOT NULL",
                           (expert_ids,))
            regions = [row[0] for row in cursor.fetchall() if row[0]]

            # Получаем рубрики экспертов
            cursor.execute("""
                SELECT DISTINCT rubric 
                FROM expert_grnti 
                WHERE id = ANY(%s) AND rubric IS NOT NULL
                ORDER BY rubric
            """, (expert_ids,))
            rubrics = [str(row[0]) for row in cursor.fetchall() if row[0] is not None]

            # Формируем строки для полей
            regions_str = ", ".join(sorted(regions)) if regions else None
            rubrics_str = ", ".join(sorted(rubrics)) if rubrics else None

            print(f"Обновление полей группы {group_id}: регионы={regions_str}, рубрики={rubrics_str}")

            # Обновляем поля группы
            cursor.execute("""
                UPDATE expert_group 
                SET regions = %s, rubric_code = %s 
                WHERE id = %s
            """, (regions_str, rubrics_str, group_id))

            print(f"Поля группы {group_id} успешно обновлены")

        except Exception as e:
            print(f"Ошибка обновления полей группы {group_id}: {e}")
        finally:
            cursor.close()

    def check_expert_dependencies(self, expert_id):
        """Проверяет, состоит ли эксперт в каких-либо группах"""
        cursor = self.connection.cursor()
        try:
            # Проверяем использование в таблице expert_group_link
            cursor.execute("""
                SELECT COUNT(*) 
                FROM expert_group_link 
                WHERE expert_id = %s
            """, (expert_id,))
            group_count = cursor.fetchone()[0]

            if group_count > 0:
                # Получаем информацию о группах, в которых состоит эксперт
                cursor.execute("""
                    SELECT g.id, g.name 
                    FROM expert_group g
                    JOIN expert_group_link l ON g.id = l.group_id
                    WHERE l.expert_id = %s
                    ORDER BY g.name
                """, (expert_id,))
                groups = cursor.fetchall()

                group_info = []
                for _, group_name in groups:
                    group_info.append(f"'{group_name}'")

                dependency_info = f"Эксперт состоит в {group_count} группе(ах):\n" + "\n".join(group_info)
                return True, dependency_info

            return False, None

        except Exception as e:
            print(f"Ошибка проверки зависимостей для эксперта {expert_id}: {e}")
            return False, None
        finally:
            cursor.close()

    def get_table_data(self, table_name, sort_column=None, sort_order='ASC'):
        """Получить данные из конкретной таблицы с возможностью сортировки"""
        cursor = self.connection.cursor()
        try:
            columns = self.get_columns_names(table_name)

            if sort_column and sort_column in columns:
                numeric_columns = {
                    'grnti_classifier': ['codrub'],
                    'expert_grnti': ['id', 'rubric'],
                    'expert': ['id', 'group_count'],
                    'reg_obl_city': ['id'],
                    'expert_group': ['id', 'participant_count']
                }

                if table_name in numeric_columns and sort_column in numeric_columns[table_name]:
                    order_by = f"CAST({sort_column} AS INTEGER) {sort_order}"
                else:
                    order_by = f'"{sort_column}" {sort_order}'
            else:
                order_by = f'{columns[0]} ASC'

            cursor.execute(f'SELECT * FROM "{table_name}" ORDER BY {order_by}')
            data = cursor.fetchall()
            return data
        except Exception as e:
            print(f"Ошибка получения данных таблицы {table_name}: {e}")
            return []
        finally:
            cursor.close()

    def get_experts_with_grnti(self, sort_column=None, sort_order='ASC'):
        """Получить данные экспертов с кодами ГРНТИ"""
        cursor = self.connection.cursor()
        try:
            base_query = """
                SELECT 
                    e.id,
                    e.name,
                    e.region,
                    e.city,
                    e.input_date,
                    e.keywords,
                    e.group_count,
                    STRING_AGG(
                        CONCAT_WS('.', 
                            g.codrub::text, 
                            eg.subrubric, 
                            eg.siscipline
                        ), 
                        '; '
                        ORDER BY g.codrub
                    ) AS grnti
                FROM expert e
                LEFT JOIN expert_grnti eg ON e.id = eg.id
                LEFT JOIN grnti_classifier g ON eg.rubric = g.codrub
                GROUP BY e.id, e.name, e.region, e.city, e.input_date, e.keywords, e.group_count
            """

            if sort_column:
                sort_mapping = {
                    'id': 'e.id',
                    'name': 'e.name',
                    'region': 'e.region',
                    'city': 'e.city',
                    'input_date': 'e.input_date',
                    'keywords': 'e.keywords',
                    'group_count': 'e.group_count',
                    'grnti': 'grnti'
                }

                numeric_columns = ['id', 'group_count']

                if sort_column in sort_mapping:
                    db_column = sort_mapping[sort_column]
                    if sort_column in numeric_columns:
                        order_by = f"CAST({db_column} AS INTEGER) {sort_order}"
                    else:
                        order_by = f"{db_column} {sort_order}"
                else:
                    order_by = "e.name ASC"
            else:
                order_by = "e.name ASC"

            query = f"{base_query} ORDER BY {order_by}"

            cursor.execute(query)
            data = cursor.fetchall()
            return data
        except Exception as e:
            print(f"Ошибка получения данных экспертов с ГРНТИ: {e}")
            return []
        finally:
            cursor.close()

    def get_joined_experts_data(self, sort_column=None, sort_order='ASC'):
        """Получить объединенные данные экспертов с их специализациями и описаниями"""
        cursor = self.connection.cursor()
        try:
            base_query = """
                SELECT 
                    e.id AS expert_id,
                    e.name AS expert_name,
                    e.region,
                    e.city,
                    e.input_date,
                    e.keywords,
                    e.group_count,
                    STRING_AGG(
                        CONCAT_WS('.', 
                            g.codrub::text, 
                            eg.subrubric, 
                            eg.siscipline
                        ), 
                        '; '
                        ORDER BY g.codrub
                    ) AS grnti,
                    STRING_AGG(
                        g.description, 
                        '; '
                        ORDER BY g.codrub
                    ) AS grnti_descriptions
                FROM expert e
                LEFT JOIN expert_grnti eg ON e.id = eg.id
                LEFT JOIN grnti_classifier g ON eg.rubric = g.codrub
                GROUP BY e.id, e.name, e.region, e.city, e.input_date, e.keywords, e.group_count
            """

            if sort_column:
                sort_mapping = {
                    'expert_id': 'e.id',
                    'expert_name': 'e.name',
                    'region': 'e.region',
                    'city': 'e.city',
                    'input_date': 'e.input_date',
                    'keywords': 'e.keywords',
                    'group_count': 'e.group_count',
                    'grnti': 'grnti',
                    'grnti_descriptions': 'grnti_descriptions'
                }

                numeric_columns = ['expert_id', 'group_count']

                if sort_column in sort_mapping:
                    db_column = sort_mapping[sort_column]
                    if sort_column in numeric_columns:
                        order_by = f"CAST({db_column} AS INTEGER) {sort_order}"
                    else:
                        order_by = f"{db_column} {sort_order}"
                else:
                    order_by = "e.name ASC"
            else:
                order_by = "e.name ASC"

            query = f"{base_query} ORDER BY {order_by}"

            cursor.execute(query)
            data = cursor.fetchall()
            return data
        except Exception as e:
            print(f"Ошибка получения объединенных данных экспертов: {e}")
            return []
        finally:
            cursor.close()

    def get_filtered_joined_experts_data(self, filters=None, sort_column=None, sort_order='ASC'):
        """Получить отфильтрованные объединенные данные экспертов с их специализациями"""
        cursor = self.connection.cursor()
        try:
            # Базовый запрос с EXISTS подзапросами для фильтрации
            base_query = """
                SELECT 
                    e.id AS expert_id,
                    e.name AS expert_name,
                    e.region,
                    e.city,
                    e.input_date,
                    e.keywords,
                    e.group_count,
                    STRING_AGG(
                        CONCAT_WS('.', 
                            g.codrub::text, 
                            eg.subrubric, 
                            eg.siscipline
                        ), 
                        '; '
                        ORDER BY g.codrub
                    ) AS grnti,
                    STRING_AGG(
                        g.description, 
                        '; '
                        ORDER BY g.codrub
                    ) AS grnti_descriptions
                FROM expert e
                LEFT JOIN expert_grnti eg ON e.id = eg.id
                LEFT JOIN grnti_classifier g ON eg.rubric = g.codrub
                WHERE 1=1
            """

            where_conditions = []
            params = []

            if filters:
                # ФИО эксперта (поддержка нескольких значений с поиском по подстроке)
                if 'expert_names' in filters:
                    name_conditions = []
                    for name in filters['expert_names']:
                        name_conditions.append("e.name ILIKE %s")
                        params.append(f"%{name}%")

                    if name_conditions:
                        where_conditions.append("(" + " OR ".join(name_conditions) + ")")

                # Регион (поддержка нескольких значений)
                if 'regions' in filters:
                    region_condition = "e.region IN ({})"
                    placeholders = ','.join(['%s'] * len(filters['regions']))
                    where_conditions.append(region_condition.format(placeholders))
                    params.extend(filters['regions'])

                # Город (поддержка нескольких значений)
                if 'cities' in filters:
                    city_condition = "e.city IN ({})"
                    placeholders = ','.join(['%s'] * len(filters['cities']))
                    where_conditions.append(city_condition.format(placeholders))
                    params.extend(filters['cities'])

                # Ключевые слова
                if 'keywords' in filters:
                    where_conditions.append("e.keywords ILIKE %s")
                    params.append(f"%{filters['keywords']}%")

                # Количество групп (новые условия с операторами)
                if 'group_conditions' in filters:
                    group_conditions = []
                    for condition in filters['group_conditions']:
                        # Разбираем условие на оператор и значение
                        parts = condition.split()
                        if len(parts) >= 2:
                            operator = parts[0]
                            value = parts[1]
                            try:
                                value_int = int(value)
                                group_conditions.append(f"e.group_count {operator} %s")
                                params.append(value_int)
                            except ValueError:
                                continue

                    if group_conditions:
                        where_conditions.append("(" + " OR ".join(group_conditions) + ")")

                # Коды ГРНТИ (поддержка нескольких значений)
                if 'grnti_codes' in filters:
                    grnti_condition = "EXISTS (SELECT 1 FROM expert_grnti eg2 WHERE eg2.id = e.id AND eg2.rubric IN ({}))"
                    placeholders = ','.join(['%s'] * len(filters['grnti_codes']))
                    where_conditions.append(grnti_condition.format(placeholders))
                    params.extend(filters['grnti_codes'])

                # Подрубрика (поддержка нескольких значений)
                if 'subrubrics' in filters:
                    subrubric_condition = "EXISTS (SELECT 1 FROM expert_grnti eg2 WHERE eg2.id = e.id AND eg2.subrubric IN ({}))"
                    placeholders = ','.join(['%s'] * len(filters['subrubrics']))
                    where_conditions.append(subrubric_condition.format(placeholders))
                    params.extend(filters['subrubrics'])

                # Дисциплина (поддержка нескольких значений)
                if 'disciplines' in filters:
                    discipline_condition = "EXISTS (SELECT 1 FROM expert_grnti eg2 WHERE eg2.id = e.id AND eg2.siscipline IN ({}))"
                    placeholders = ','.join(['%s'] * len(filters['disciplines']))
                    where_conditions.append(discipline_condition.format(placeholders))
                    params.extend(filters['disciplines'])

            # Формируем полный запрос
            if where_conditions:
                base_query += " AND " + " AND ".join(where_conditions)

            base_query += " GROUP BY e.id, e.name, e.region, e.city, e.input_date, e.keywords, e.group_count"

            # Добавляем сортировку если указана
            if sort_column:
                sort_mapping = {
                    'expert_id': 'e.id',
                    'expert_name': 'e.name',
                    'region': 'e.region',
                    'city': 'e.city',
                    'input_date': 'e.input_date',
                    'keywords': 'e.keywords',
                    'group_count': 'e.group_count',
                    'grnti': 'grnti',
                    'grnti_descriptions': 'grnti_descriptions'
                }

                numeric_columns = ['expert_id', 'group_count']

                if sort_column in sort_mapping:
                    db_column = sort_mapping[sort_column]
                    if sort_column in numeric_columns:
                        order_by = f"CAST({db_column} AS INTEGER) {sort_order}"
                    else:
                        order_by = f"{db_column} {sort_order}"
                else:
                    order_by = "e.name ASC"
            else:
                order_by = "e.name ASC"

            query = f"{base_query} ORDER BY {order_by}"

            cursor.execute(query, params)
            data = cursor.fetchall()
            return data
        except Exception as e:
            print(f"Ошибка выполнения запроса фильтрации: {e}")
            return []
        finally:
            cursor.close()

    def get_columns_names(self, table_name):
        """Получить названия столбцов таблицы"""
        cursor = self.connection.cursor()
        try:
            cursor.execute(f"""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = '{table_name}'
                ORDER BY ordinal_position
            """)
            columns = [row[0] for row in cursor.fetchall()]
            return columns
        except Exception as e:
            print(f"Ошибка получения названий столбцов для {table_name}: {e}")
            return []
        finally:
            cursor.close()

    def insert_record(self, table_name, data):
        """Добавить новую запись в таблицу"""
        cursor = self.connection.cursor()
        try:
            columns = self.get_columns_names(table_name)

            placeholders = ', '.join(['%s'] * len(data))
            columns_str = ', '.join(columns)

            query = f"INSERT INTO {table_name} ({columns_str}) VALUES ({placeholders})"

            cursor.execute(query, data)
            self.connection.commit()
            return True
        except Exception as e:
            self.connection.rollback()
            raise e
        finally:
            cursor.close()

    def insert_expert_record(self, data):
        """Специальный метод для вставки эксперта с правильным порядком полей"""
        cursor = self.connection.cursor()
        try:
            query = """
                INSERT INTO expert (id, name, region, city, input_date, keywords, group_count) 
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """

            cursor.execute(query, data)
            self.connection.commit()
            return True
        except Exception as e:
            self.connection.rollback()
            raise e
        finally:
            cursor.close()

    def update_record(self, table_name, record_id, data):
        """Обновить запись в таблице"""
        cursor = self.connection.cursor()
        try:
            columns = self.get_columns_names(table_name)

            set_clause = ', '.join([f"{col} = %s" for col in columns[1:]])
            query = f"UPDATE {table_name} SET {set_clause} WHERE {columns[0]} = %s"

            data_with_id = data + [record_id]

            cursor.execute(query, data_with_id)
            self.connection.commit()
        except Exception as e:
            self.connection.rollback()
            raise e
        finally:
            cursor.close()

    def update_expert_record(self, record_id, data):
        """Специальный метод для обновления эксперта с правильным порядком полей"""
        cursor = self.connection.cursor()
        try:
            query = """
                UPDATE expert 
                SET name = %s, region = %s, city = %s, input_date = %s, keywords = %s, group_count = %s
                WHERE id = %s
            """

            data_with_id = data + [record_id]

            cursor.execute(query, data_with_id)
            self.connection.commit()
        except Exception as e:
            self.connection.rollback()
            raise e
        finally:
            cursor.close()

    def delete_record(self, table_name, record_id):
        """Удалить запись из таблицы"""
        cursor = self.connection.cursor()
        try:
            columns = self.get_columns_names(table_name)

            query = f"DELETE FROM {table_name} WHERE {columns[0]} = %s"

            cursor.execute(query, (record_id,))
            self.connection.commit()
        except Exception as e:
            self.connection.rollback()
            raise e
        finally:
            cursor.close()

    def get_regions(self):
        """Получить список уникальных регионов"""
        cursor = self.connection.cursor()
        try:
            cursor.execute("SELECT DISTINCT region FROM reg_obl_city ORDER BY region")
            regions = [row[0] for row in cursor.fetchall()]
            return regions
        except Exception as e:
            print(f"Ошибка получения регионов: {e}")
            return []
        finally:
            cursor.close()

    def get_cities_by_region(self, region):
        """Получить список городов по региону"""
        cursor = self.connection.cursor()
        try:
            cursor.execute("SELECT DISTINCT city FROM reg_obl_city WHERE region = %s ORDER BY city", (region,))
            cities = [row[0] for row in cursor.fetchall()]
            return cities
        except Exception as e:
            print(f"Ошибка получения городов для региона {region}: {e}")
            return []
        finally:
            cursor.close()

    def get_all_cities_with_regions(self):
        """Получить все города с их регионами"""
        cursor = self.connection.cursor()
        try:
            cursor.execute("SELECT city, region FROM reg_obl_city ORDER BY city")
            data = cursor.fetchall()
            return data
        except Exception as e:
            print(f"Ошибка получения всех городов: {e}")
            return []
        finally:
            cursor.close()

    def get_expert_grnti(self, expert_id):
        """Получить коды ГРНТИ для эксперта"""
        cursor = self.connection.cursor()
        try:
            cursor.execute("SELECT rubric, subrubric, siscipline FROM expert_grnti WHERE id = %s", (expert_id,))
            data = cursor.fetchall()
            return data
        except Exception as e:
            print(f"Ошибка получения кодов ГРНТИ для эксперта {expert_id}: {e}")
            return []
        finally:
            cursor.close()

    def save_expert_grnti(self, expert_id, grnti_data):
        """Сохранить коды ГРНТИ для эксперта"""
        cursor = self.connection.cursor()
        try:
            cursor.execute("DELETE FROM expert_grnti WHERE id = %s", (expert_id,))

            for code, subrubric, discipline in grnti_data:
                try:
                    code_int = int(code) if code else 0
                except (ValueError, TypeError):
                    code_int = 0

                subrubric_val = subrubric if subrubric else None
                discipline_val = discipline if discipline else None

                cursor.execute(
                    "INSERT INTO expert_grnti (id, rubric, subrubric, siscipline) VALUES (%s, %s, %s, %s)",
                    (expert_id, code_int, subrubric_val, discipline_val)
                )

            self.connection.commit()
        except Exception as e:
            self.connection.rollback()
            raise e
        finally:
            cursor.close()

    def get_group_members_with_details(self, group_id):
        """Получить участников группы с регионами и рубриками ГРНТИ"""
        cursor = self.connection.cursor()
        try:
            cursor.execute("""
                SELECT e.id, e.name, e.region, eg.rubric
                FROM expert_group_link l
                JOIN expert e ON l.expert_id = e.id
                LEFT JOIN expert_grnti eg ON e.id = eg.id
                WHERE l.group_id = %s
                ORDER BY e.name
            """, (group_id,))
            return cursor.fetchall()  # [(id, name, region, rubric), ...]
        finally:
            cursor.close()

    def check_grnti_dependencies(self, grnti_code):
        """Проверяет, используется ли код ГРНТИ в других таблицах"""
        cursor = self.connection.cursor()
        try:
            # Проверяем использование в таблице expert_grnti
            cursor.execute("SELECT COUNT(*) FROM expert_grnti WHERE rubric = %s", (grnti_code,))
            count = cursor.fetchone()[0]
            return count > 0
        except Exception as e:
            print(f"Ошибка проверки зависимостей для ГРНТИ {grnti_code}: {e}")
            return False
        finally:
            cursor.close()

    def check_region_city_dependencies(self, record_id):
        """Проверяет, используется ли регион или город в таблице экспертов"""
        cursor = self.connection.cursor()
        try:
            # Получаем данные региона/города по ID
            cursor.execute("SELECT region, city FROM reg_obl_city WHERE id = %s", (record_id,))
            result = cursor.fetchone()

            if not result:
                return False, None

            region, city = result

            # Проверяем использование региона
            cursor.execute("SELECT COUNT(*) FROM expert WHERE region = %s", (region,))
            region_count = cursor.fetchone()[0]

            # Проверяем использование города
            cursor.execute("SELECT COUNT(*) FROM expert WHERE city = %s", (city,))
            city_count = cursor.fetchone()[0]

            has_dependencies = region_count > 0 or city_count > 0

            dependency_info = []
            if region_count > 0:
                dependency_info.append(f"регион '{region}' используется {region_count} раз(а)")
            if city_count > 0:
                dependency_info.append(f"город '{city}' используется {city_count} раз(а)")

            return has_dependencies, ", ".join(dependency_info)

        except Exception as e:
            print(f"Ошибка проверки зависимостей для региона/города {record_id}: {e}")
            return False, None
        finally:
            cursor.close()

    def get_unique_rubrics(self):
        """Получить уникальные рубрики (только rubric из expert_grnti)"""
        cursor = self.connection.cursor()
        try:
            cursor.execute("SELECT DISTINCT rubric FROM expert_grnti ORDER BY rubric")
            rubrics = [row[0] for row in cursor.fetchall() if row[0] is not None]
            return rubrics
        finally:
            cursor.close()

    def get_experts_for_group(self, region=None, rubric=None):
        cursor = self.connection.cursor()
        try:
            # Используем подзапрос или EXISTS, чтобы избежать дубликатов
            base_query = """
                SELECT e.id, e.name, e.region,
                    (SELECT STRING_AGG(DISTINCT rubric::text, ', ')
                        FROM expert_grnti eg2
                        WHERE eg2.id = e.id
                    ) AS rubrics
                FROM expert e
                WHERE 1=1
            """
            conditions = []
            params = []

            if rubric is not None:
                # Фильтруем по наличию рубрики у эксперта
                base_query += " AND EXISTS (SELECT 1 FROM expert_grnti eg WHERE eg.id = e.id AND eg.rubric = %s)"
                params.append(rubric)

            if region is not None:
                conditions.append("e.region = %s")
                params.append(region)

            if conditions:
                base_query += " AND " + " AND ".join(conditions)

            base_query += " ORDER BY e.name"

            cursor.execute(base_query, params)
            results = cursor.fetchall()

            # Преобразуем: для совместимости с существующим кодом возвращаем (id, name, region, rubric)
            # Берём первую рубрику из списка (или None, если нет)
            output = []
            for row in results:
                eid, name, reg, rubrics_str = row
                rubric_val = None
                if rubrics_str:
                    # Берём первую рубрику (можно и любую — главное, чтобы не дублировать эксперта)
                    rubric_val = int(rubrics_str.split(',')[0].strip())
                output.append((eid, name, reg, rubric_val))

            return output
        finally:
            cursor.close()

    def get_group_members(self, group_id):
        cursor = self.connection.cursor()
        try:
            cursor.execute("""
                SELECT e.id, e.name
                FROM expert_group_link l
                JOIN expert e ON l.expert_id = e.id
                WHERE l.group_id = %s
                ORDER BY e.name
            """, (group_id,))
            return cursor.fetchall()
        finally:
            cursor.close()

    def get_grnti_details_for_experts(self, expert_ids):
        """Возвращает словарь с подробной информацией по ГРНТИ для указанных экспертов"""
        if not expert_ids:
            return {}

        expert_ids = list({eid for eid in expert_ids if eid is not None})
        if not expert_ids:
            return {}

        cursor = self.connection.cursor()
        try:
            cursor.execute("""
                SELECT 
                    eg.id AS expert_id,
                    g.codrub::text AS base_code,
                    eg.subrubric,
                    eg.siscipline,
                    g.description
                FROM expert_grnti eg
                LEFT JOIN grnti_classifier g ON eg.rubric = g.codrub
                WHERE eg.id = ANY(%s)
                ORDER BY g.codrub, eg.subrubric, eg.siscipline
            """, (expert_ids,))

            details_map = {}
            for expert_id, base_code, subrubric, discipline, description in cursor.fetchall():
                entries = details_map.setdefault(expert_id, [])
                code_parts = []
                if base_code:
                    code_parts.append(str(base_code))
                if subrubric:
                    code_parts.append(str(subrubric))
                if discipline:
                    code_parts.append(str(discipline))
                full_code = ".".join(code_parts) if code_parts else ""
                entries.append({
                    'base_code': base_code,
                    'subrubric': subrubric,
                    'discipline': discipline,
                    'description': description,
                    'code_full': full_code
                })

            return details_map
        finally:
            cursor.close()

    def insert_record_without_id(self, table_name, data):
        """Вставляет запись, пропуская первый столбец (обычно id)"""
        cursor = self.connection.cursor()
        try:
            columns = self.get_columns_names(table_name)
            # Пропускаем первый столбец (id)
            columns_no_id = columns[1:]
            placeholders = ', '.join(['%s'] * len(data))
            columns_str = ', '.join(columns_no_id)
            query = f"INSERT INTO {table_name} ({columns_str}) VALUES ({placeholders})"
            cursor.execute(query, data)
            self.connection.commit()
            return True
        except Exception as e:
            self.connection.rollback()
            raise e
        finally:
            cursor.close()


class ExpertEditDialog(QDialog):
    """Специальное диалоговое окно для добавления/редактирования экспертов"""

    def __init__(self, table_name, columns, data=None, parent=None, db=None, is_edit=False, expert_id=None):
        super().__init__(parent)
        self.table_name = table_name
        self.columns = columns
        self.data = data
        self.db = db
        self.is_edit = is_edit
        self.expert_id = expert_id
        self.grnti_data = []
        self.all_cities = []

        self.setup_ui()

    def setup_ui(self):
        if self.is_edit:
            self.setWindowTitle("Редактирование записи об эксперте")
        else:
            self.setWindowTitle("Добавление записи об эксперте")

        self.setMinimumSize(500, 400)

        self.layout = QVBoxLayout()

        # Поля для ввода ФИО - разные для создания и редактирования
        if self.is_edit:
            # При редактировании - одно поле ФИО
            name_label = QLabel("ФИО эксперта*")
            name_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
            self.layout.addWidget(name_label)

            self.name_field = QLineEdit()
            self.name_field.setPlaceholderText("Например: Иванов И.И. или Петров П.")
            self.name_field.setStyleSheet("""
                QLineEdit {
                    padding: 8px;
                    border: 1px solid #bdc3c7;
                    border-radius: 4px;
                    background-color: white;
                    margin-bottom: 10px;
                }
                QLineEdit:focus {
                    border-color: #3498db;
                }
            """)
            self.layout.addWidget(self.name_field)
            
            # Для редактирования не создаем отдельные поля
            self.surname_field = None
            self.patronymic_field = None
        else:
            # При создании - три отдельных поля
            # Фамилия
            surname_label = QLabel("Фамилия*")
            surname_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
            self.layout.addWidget(surname_label)
            
            self.surname_field = QLineEdit()
            self.surname_field.setPlaceholderText("Например: Иванов")
            self.surname_field.setStyleSheet("""
                QLineEdit {
                    padding: 8px;
                    border: 1px solid #bdc3c7;
                    border-radius: 4px;
                    background-color: white;
                    margin-bottom: 10px;
                }
                QLineEdit:focus {
                    border-color: #3498db;
                }
            """)
            self.layout.addWidget(self.surname_field)
            
            # Имя
            name_label = QLabel("Имя*")
            name_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
            self.layout.addWidget(name_label)
            
            self.name_field = QLineEdit()
            self.name_field.setPlaceholderText("Например: Иван")
            self.name_field.setStyleSheet("""
                QLineEdit {
                    padding: 8px;
                    border: 1px solid #bdc3c7;
                    border-radius: 4px;
                    background-color: white;
                    margin-bottom: 10px;
                }
                QLineEdit:focus {
                    border-color: #3498db;
                }
            """)
            self.layout.addWidget(self.name_field)
            
            # Отчество (необязательное)
            patronymic_label = QLabel("Отчество")
            patronymic_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
            self.layout.addWidget(patronymic_label)
            
            self.patronymic_field = QLineEdit()
            self.patronymic_field.setPlaceholderText("Например: Иванович")
            self.patronymic_field.setStyleSheet("""
                QLineEdit {
                    padding: 8px;
                    border: 1px solid #bdc3c7;
                    border-radius: 4px;
                    background-color: white;
                    margin-bottom: 10px;
                }
                QLineEdit:focus {
                    border-color: #3498db;
                }
            """)
            self.layout.addWidget(self.patronymic_field)

        # Выбор региона
        region_label = QLabel("Регион*")
        region_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        self.layout.addWidget(region_label)

        self.region_combo = QComboBox()
        self.region_combo.addItem("")
        regions = self.db.get_regions() if self.db else []
        for region in regions:
            self.region_combo.addItem(region)

        self.region_combo.currentTextChanged.connect(self.on_region_changed)
        self.region_combo.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
                margin-bottom: 10px;
            }
            QComboBox:focus {
                border-color: #3498db;
            }
        """)
        self.layout.addWidget(self.region_combo)

        # Выбор города с автодополнением
        city_label = QLabel("Город*")
        city_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        self.layout.addWidget(city_label)

        self.city_combo = QComboBox()
        self.city_combo.setEditable(True)

        self.all_cities = self.db.get_all_cities_with_regions() if self.db else []
        city_names = [city for city, region in self.all_cities]
        self.city_combo.addItems([""] + city_names)

        # Настраиваем автодополнение
        completer = QCompleter(city_names)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        self.city_combo.setCompleter(completer)

        self.city_combo.currentTextChanged.connect(self.on_city_changed)
        self.city_combo.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
                margin-bottom: 10px;
            }
            QComboBox:focus {
                border-color: #3498db;
            }
        """)
        self.layout.addWidget(self.city_combo)

        # Ключевые слова
        keywords_label = QLabel("Ключевые слова")
        keywords_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        self.layout.addWidget(keywords_label)

        self.keywords_field = QLineEdit()
        self.keywords_field.setPlaceholderText("Введите ключевые слова через запятую")
        self.keywords_field.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
                margin-bottom: 10px;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
        """)
        self.layout.addWidget(self.keywords_field)

        # Кнопка для управления кодами ГРНТИ
        grnti_layout = QHBoxLayout()
        grnti_label = QLabel("Коды ГРНТИ:")
        grnti_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        self.grnti_button = QPushButton("Управление кодами ГРНТИ")
        self.grnti_button.clicked.connect(self.manage_grnti)
        self.grnti_button.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
            QPushButton:pressed {
                background-color: #7d3c98;
            }
        """)
        grnti_layout.addWidget(grnti_label)
        grnti_layout.addWidget(self.grnti_button)
        self.layout.addLayout(grnti_layout)

        # Дата добавления
        date_label = QLabel("Дата добавления")
        date_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        self.layout.addWidget(date_label)

        self.date_field = QLineEdit()
        self.date_field.setReadOnly(True)
        self.date_field.setText(datetime.now().strftime('%d.%m.%Y'))
        self.date_field.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: #ecf0f1;
                margin-bottom: 15px;
                color: #7f8c8d;
            }
        """)
        self.layout.addWidget(self.date_field)

        # Заполняем данные, если это редактирование
        if self.data:
            self.fill_form_data()

        # Загружаем коды ГРНТИ, если это редактирование и есть expert_id
        if self.is_edit and self.expert_id and self.db:
            self.grnti_data = self.db.get_expert_grnti(self.expert_id)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                                   QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.validate_and_accept)
        buttons.rejected.connect(self.reject)

        # Изменяем текст кнопок на русский
        ok_button = buttons.button(QDialogButtonBox.StandardButton.Ok)
        ok_button.setText("ОК")
        cancel_button = buttons.button(QDialogButtonBox.StandardButton.Cancel)
        cancel_button.setText("Отмена")

        # Стилизация кнопок диалога
        for button in buttons.buttons():
            if buttons.buttonRole(button) == QDialogButtonBox.ButtonRole.AcceptRole:
                button.setStyleSheet("""
                    QPushButton {
                        background-color: #27ae60;
                        color: white;
                        border: none;
                        padding: 8px 16px;
                        border-radius: 4px;
                        font-weight: bold;
                        min-width: 80px;
                    }
                    QPushButton:hover {
                        background-color: #219a52;
                    }
                    QPushButton:pressed {
                        background-color: #1e8449;
                    }
                """)
            else:
                button.setStyleSheet("""
                    QPushButton {
                        background-color: #95a5a6;
                        color: white;
                        border: none;
                        padding: 8px 16px;
                        border-radius: 4px;
                        font-weight: bold;
                        min-width: 80px;
                    }
                    QPushButton:hover {
                        background-color: #7f8c8d;
                    }
                    QPushButton:pressed {
                        background-color: #707b7c;
                    }
                """)

        self.layout.addWidget(buttons)

        self.setLayout(self.layout)

    def on_region_changed(self, region):
        """Обновляет список городов при изменении региона"""
        current_city = self.city_combo.currentText()
        self.city_combo.clear()
        self.city_combo.addItem("")
        if region and self.db:
            cities = self.db.get_cities_by_region(region)
            for city in cities:
                self.city_combo.addItem(city)

        # Восстанавливаем предыдущее значение города, если оно соответствует новому региону
        if current_city and self.all_cities:
            for city_name, city_region in self.all_cities:
                if city_name == current_city and city_region == region:
                    self.city_combo.setCurrentText(current_city)
                    break

    def on_city_changed(self, city):
        """Автоматически устанавливает регион при выборе города"""
        if city and self.all_cities:
            for city_name, region in self.all_cities:
                if city_name == city:
                    self.region_combo.blockSignals(True)
                    self.region_combo.setCurrentText(region)
                    self.region_combo.blockSignals(False)
                    break

    def fill_form_data(self):
        """Заполняет форму данными при редактировании"""
        if self.data:
            if len(self.data) >= 1:
                name = str(self.data[0]) if self.data[0] is not None else ""
                if self.is_edit:
                    # При редактировании просто заполняем одно поле ФИО
                    self.name_field.setText(name)
                else:
                    # При создании разбираем ФИО на части
                    name_parts = self.parse_name(name)
                    if self.surname_field:
                        self.surname_field.setText(name_parts.get('surname', ''))
                    if self.name_field:
                        self.name_field.setText(name_parts.get('name', ''))
                    if self.patronymic_field:
                        self.patronymic_field.setText(name_parts.get('patronymic', ''))

            if len(self.data) >= 2:
                region = str(self.data[1]) if self.data[1] is not None else ""
                self.region_combo.setCurrentText(region)
                if region and self.db:
                    self.on_region_changed(region)

            if len(self.data) >= 3:
                city = str(self.data[2]) if self.data[2] is not None else ""
                self.city_combo.setCurrentText(city)

            if len(self.data) >= 4:
                date_value = str(self.data[3]) if self.data[3] is not None else ""
                if date_value:
                    try:
                        if hasattr(self.data[3], 'strftime'):
                            formatted_date = self.data[3].strftime('%d.%m.%Y')
                        else:
                            date_obj = DateValidator.parse_date(date_value)
                            if date_obj:
                                formatted_date = date_obj.strftime('%d.%m.%Y')
                            else:
                                formatted_date = date_value
                        self.date_field.setText(formatted_date)
                    except Exception:
                        self.date_field.setText(date_value)

            if len(self.data) >= 5:
                keywords = str(self.data[4]) if self.data[4] is not None else ""
                self.keywords_field.setText(keywords)

    def manage_grnti(self):
        """Открывает диалог управления кодами ГРНТИ"""
        expert_id = self.expert_id

        if expert_id and self.db:
            grnti_data = self.db.get_expert_grnti(expert_id)
            self.grnti_data = grnti_data

        dialog = GRNTIDialog(self, self.db, expert_id, self.grnti_data)
        if dialog.exec():
            self.grnti_data = dialog.get_grnti_data()

    def validate_and_accept(self):
        """Проверяет валидность данных перед принятием"""
        if self.is_edit:
            # При редактировании проверяем одно поле ФИО
            name = self.name_field.text().strip()
            if not name:
                QMessageBox.warning(self, "Ошибка", "Поле 'ФИО эксперта' обязательно для заполнения")
                self.name_field.setFocus()
                return
        else:
            # При создании проверяем отдельные поля
            surname = self.surname_field.text().strip()
            name = self.name_field.text().strip()
            
            if not surname:
                QMessageBox.warning(self, "Ошибка", "Поле 'Фамилия' обязательно для заполнения")
                self.surname_field.setFocus()
                return
            
            if not name:
                QMessageBox.warning(self, "Ошибка", "Поле 'Имя' обязательно для заполнения")
                self.name_field.setFocus()
                return

        if not self.region_combo.currentText().strip():
            QMessageBox.warning(self, "Ошибка", "Поле 'Регион' обязательно для заполнения")
            self.region_combo.setFocus()
            return

        if not self.city_combo.currentText().strip():
            QMessageBox.warning(self, "Ошибка", "Поле 'Город' обязательно для заполнения")
            self.city_combo.setFocus()
            return

        if not self.grnti_data:
            QMessageBox.warning(self, "Ошибка", "Добавьте хотя бы один код ГРНТИ")
            return

        self.accept()
    
    def parse_name(self, name_str):
        """Разбирает ФИО в формате 'Фамилия И.О.' или 'Фамилия И.' на отдельные части"""
        result = {'surname': '', 'name': '', 'patronymic': ''}
        
        if not name_str:
            return result
        
        name_str = name_str.strip()
        parts = name_str.split()
        if len(parts) == 0:
            return result
        
        # Первая часть - всегда фамилия
        result['surname'] = parts[0]
        
        # Остальные части - инициалы или полные имена
        if len(parts) > 1:
            second_part = parts[1]
            
            # Если формат "И.О." (например, "И.И." - две буквы с точкой между ними)
            if '.' in second_part:
                # Убираем точки и разбиваем
                # Может быть "И.О." или "И.И." - одна строка с точкой посередине
                if second_part.count('.') == 1 and len(second_part) == 3:
                    # Формат "И.О" (например, "И.И")
                    initials = second_part.split('.')
                    if len(initials) >= 1 and initials[0]:
                        result['name'] = initials[0]
                    if len(initials) >= 2 and initials[1]:
                        result['patronymic'] = initials[1]
                elif len(parts) > 2:
                    # Формат "И." "О." (отдельные части)
                    result['name'] = parts[1].replace('.', '')
                    result['patronymic'] = parts[2].replace('.', '')
                else:
                    # Только "И."
                    result['name'] = second_part.replace('.', '')
            else:
                # Полное имя (без точки)
                result['name'] = second_part
                if len(parts) > 2:
                    result['patronymic'] = parts[2]
        
        return result

    def get_data(self):
        """Получить данные из полей ввода"""
        if self.is_edit:
            # При редактировании просто берем ФИО из одного поля
            formatted_name = self.name_field.text().strip()
        else:
            # При создании формируем ФИО из трех полей с автоматическим сокращением
            surname = self.surname_field.text().strip()
            name = self.name_field.text().strip()
            patronymic = self.patronymic_field.text().strip()
            
            # Берем первую букву имени и отчества
            name_initial = name[0].upper() if name else ''
            patronymic_initial = patronymic[0].upper() if patronymic else ''
            
            # Формируем ФИО
            if patronymic:
                formatted_name = f"{surname} {name_initial}.{patronymic_initial}."
            else:
                formatted_name = f"{surname} {name_initial}."
        
        region = self.region_combo.currentText().strip()
        city = self.city_combo.currentText().strip()

        if self.is_edit and self.data and len(self.data) >= 4 and self.data[3]:
            input_date = self.data[3]
            if isinstance(input_date, str):
                date_obj = DateValidator.parse_date(input_date)
                if date_obj:
                    input_date = date_obj.strftime('%Y-%m-%d')
                else:
                    input_date = datetime.now().strftime('%Y-%m-%d')
        else:
            input_date = datetime.now().strftime('%Y-%m-%d')

        keywords = self.keywords_field.text().strip()
        if not keywords:
            keywords = None

        # Количество групп всегда устанавливается в 0 при создании/редактировании эксперта
        group_count = 0

        return [formatted_name, region, city, input_date, keywords, group_count]

    def get_grnti_data(self):
        """Возвращает данные о кодах ГРНТИ"""
        return self.grnti_data

    def set_grnti_data(self, grnti_data):
        """Устанавливает данные о кодах ГРНТИ"""
        self.grnti_data = grnti_data


class SortableTableWidget:
    """Миксин для добавления функциональности сортировки таблицы"""

    def setup_sorting(self):
        """Настройка сортировки таблицы"""
        self.table_widget.setSortingEnabled(False)
        self.table_widget.horizontalHeader().sectionClicked.connect(self.on_header_clicked)

        self.current_sort_column = None
        self.current_sort_order = Qt.SortOrder.AscendingOrder

        self.sort_settings = {}

    def on_header_clicked(self, logical_index):
        """Обработчик клика по заголовку столбца"""
        if not self.current_table:
            return

        columns = self.get_columns_for_table()
        display_columns = self.get_display_columns(self.current_table, columns)

        if logical_index >= len(display_columns):
            return

        column_name = self.get_db_column_name(self.current_table, logical_index)
        if not column_name:
            return

        if self.current_sort_column == column_name:
            if self.current_sort_order == Qt.SortOrder.AscendingOrder:
                self.current_sort_order = Qt.SortOrder.DescendingOrder
                sort_order_str = 'DESC'
            else:
                self.current_sort_order = Qt.SortOrder.AscendingOrder
                sort_order_str = 'ASC'
        else:
            self.current_sort_column = column_name
            self.current_sort_order = Qt.SortOrder.AscendingOrder
            sort_order_str = 'ASC'

        self.sort_settings[self.current_table] = (column_name, sort_order_str)

        self.show_table(self.current_table, column_name, sort_order_str)

    def get_columns_for_table(self):
        """Получить столбцы для текущей таблицы"""
        if self.current_table == 'joined_experts':
            # Для joined_experts SQL возвращает: expert_id, expert_name, region, city, input_date, keywords, group_count, grnti, grnti_descriptions
            return ['expert_id', 'expert_name', 'region', 'city', 'input_date', 'keywords', 'group_count', 'grnti',
                    'grnti_descriptions']
        else:
            return self.db.get_columns_names(self.current_table)

    def get_display_columns(self, table_name, columns):
        """Получить отображаемые столбцы (исключая скрытые)"""
        display_columns = []
        for col in columns:
            if self.should_hide_column(table_name, col):
                continue
            display_columns.append(col)
        return display_columns

    def should_hide_column(self, table_name, column_name):
        """Определяет, должен ли столбец быть скрыт от пользователя"""
        column_lower = column_name.lower()
        if column_lower == 'id':
            return True
        if column_lower.endswith('_id'):
            return True
        return False

    def get_db_column_name(self, table_name, display_index):
        """Получить имя столбца в БД по индексу отображаемого столбца"""
        columns = self.get_columns_for_table()
        display_columns = self.get_display_columns(table_name, columns)

        if display_index < len(display_columns):
            return display_columns[display_index]
        return None


class MainWindow(QMainWindow, Ui_MainWindow, SortableTableWidget):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.setWindowTitle("Управление экспертизой проектов")
        self.setMinimumSize(800, 600)

        self.setStyleSheet("""
            QMainWindow {
                background-color: #ecf0f1;
            }
            QMenuBar {
                background-color: #2c3e50;
                color: white;
                padding: 5px;
            }
            QMenuBar::item {
                background-color: transparent;
                padding: 5px 10px;
            }
            QMenuBar::item:selected {
                background-color: #34495e;
            }
            QMenu {
                background-color: white;
                border: 1px solid #bdc3c7;
            }
            QMenu::item {
                padding: 5px 20px;
            }
            QMenu::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)

        self.setup_table_behavior()
        self.setup_visual_feedback()
        self.setup_sorting()

        screen = QApplication.primaryScreen()
        screen_geometry = screen.availableGeometry()
        self.resize(int(screen_geometry.width() * 0.8), int(screen_geometry.height() * 0.7))

        self.column_display_names = {
            'expert': {
                'name': 'ФИО эксперта',
                'region': 'Регион',
                'city': 'Город',
                'input_date': 'Дата',
                'keywords': 'Ключ. слова',
                'group_count': 'Кол-во групп',
                'grnti': 'ГРНТИ'
            },
            'grnti_classifier': {
                'codrub': 'Код ГРНТИ',
                'description': 'Название рубрики'
            },
            'reg_obl_city': {
                'region': 'Федеральный округ',
                'oblname': 'Субъект федерации',
                'city': 'Город'
            },
            'expert_grnti': {
                'rubric': 'Рубрика',
                'subrubric': 'Сабрубрика',
                'siscipline': 'Дисциплина'
            },
            'joined_experts': {
                'expert_name': 'ФИО эксперта',
                'region': 'Регион',
                'city': 'Город',
                'input_date': 'Дата',
                'keywords': 'Ключ. слова',
                'group_count': 'Кол-во групп',
                'grnti': 'ГРНТИ',
                'grnti_descriptions': 'Описание ГРНТИ'
            }
        }

        self.table_display_names = {
            'expert': 'Эксперты',
            'grnti_classifier': 'Классификатор ГРНТИ',
            'reg_obl_city': 'Регионы и города',
            'expert_grnti': 'Связь экспертов и ГРНТИ',
            'joined_experts': 'Общая таблица экспертов'
        }

        self.date_columns = {
            'expert': ['input_date'],
            'joined_experts': ['input_date']
        }

        self.numeric_columns = {
            'expert': ['group_count'],
            'grnti_classifier': ['codrub'],
            'expert_grnti': ['rubric'],
            'reg_obl_city': []
        }

        self.column_display_names['expert_group'] = {
            'name': 'Название',
            'participant_count': 'Кол-во участников',
            'regions': 'Регионы',
            'rubric_code': 'Код рубрики',
            'created_at': 'Дата создания'
        }

        self.table_display_names['expert_group'] = 'Группы'
        self.date_columns['expert_group'] = ['created_at']
        self.numeric_columns['expert_group'] = ['participant_count']

        try:
            self.db = DatabaseManager()
            self.connect_menu_actions()
            self.connect_button_actions()
            self.statusbar.showMessage("Подключение к базе данных успешно")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось подключиться к базе данных: {str(e)}")
            self.close()

        self.current_table = None
        self.current_table_data = {}
        self.current_filters = {}

        self.setup_filter_button()
        self.setup_view_members_button()  # НОВАЯ КНОПКА ПРОСМОТРА СОСТАВА ГРУППЫ
        self.update_button_visibility()

        self.last_added_id = None

        # Включаем контекстное меню для таблицы - ИСПРАВЛЕННАЯ РЕАЛИЗАЦИЯ
        self.table_widget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_widget.customContextMenuRequested.connect(self.show_context_menu)

    def setup_view_members_button(self):
        """Настройка кнопки просмотра состава группы"""
        self.viewMembersButton = QPushButton("Просмотр состава")
        self.viewMembersButton.clicked.connect(self.view_selected_group_members)
        self.viewMembersButton.setStyleSheet("""
            QPushButton { 
                background-color: #9b59b6; 
                color: white; 
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
            QPushButton:pressed {
                background-color: #7d3c98;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
                color: #7f8c8d;
            }
        """)

        # Добавляем кнопку в основной layout кнопок
        self.buttons_layout.addWidget(self.viewMembersButton)

    def view_selected_group_members(self):
        """Просмотр состава выбранной группы"""
        if self.current_table != 'expert_group':
            return

        selected_row = self.table_widget.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите группу для просмотра состава")
            return

        try:
            # Получаем данные выбранной группы
            if selected_row < len(self.current_table_data.get('expert_group', [])):
                group_data = self.current_table_data['expert_group'][selected_row]
                group_id = group_data[0]
                group_name = group_data[1]
                created_at = group_data[3] if len(group_data) > 3 else None

                # Открываем диалог просмотра состава группы
                dialog = GroupMembersDialog(self, self.db, group_id, group_name, created_at)
                dialog.exec()
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось получить данные выбранной группы")
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось открыть состав группы: {str(e)}")

    def update_group_fields(self, group_id):
        """Обновляет поля регионов и рубрик для группы"""
        try:
            cursor = self.db.connection.cursor()

            # Получаем всех экспертов группы
            cursor.execute("SELECT expert_id FROM expert_group_link WHERE group_id = %s", (group_id,))
            expert_ids = [row[0] for row in cursor.fetchall()]

            if not expert_ids:
                return

            # Получаем регионы экспертов
            cursor.execute("SELECT DISTINCT region FROM expert WHERE id = ANY(%s) AND region IS NOT NULL",
                           (expert_ids,))
            regions = [row[0] for row in cursor.fetchall() if row[0]]

            # Получаем рубрики экспертов
            cursor.execute("""
                SELECT DISTINCT rubric 
                FROM expert_grnti 
                WHERE id = ANY(%s) AND rubric IS NOT NULL
                ORDER BY rubric
            """, (expert_ids,))
            rubrics = [str(row[0]) for row in cursor.fetchall() if row[0] is not None]

            # Формируем строки для полей
            regions_str = ", ".join(sorted(regions)) if regions else None
            rubrics_str = ", ".join(sorted(rubrics)) if rubrics else None

            print(f"Обновление группы {group_id}: регионы={regions_str}, рубрики={rubrics_str}")

            # Обновляем поля группы
            cursor.execute("""
                UPDATE expert_group 
                SET regions = %s, rubric_code = %s 
                WHERE id = %s
            """, (regions_str, rubrics_str, group_id))

            self.db.connection.commit()
            print(f"Поля группы {group_id} успешно обновлены")

        except Exception as e:
            self.db.connection.rollback()
            print(f"Ошибка обновления полей группы {group_id}: {e}")
        finally:
            cursor.close()

    def setup_filter_button(self):
        """Настройка кнопки фильтрации"""
        self.filterButton = QPushButton("Фильтр")
        self.filterButton.clicked.connect(self.open_filter_dialog)
        self.filterButton.setStyleSheet("""
            QPushButton { 
                background-color: #3498db; 
                color: white; 
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
        """)

        self.resetFilterButton = QPushButton("Сбросить фильтры")
        self.resetFilterButton.clicked.connect(self.reset_filters)
        self.resetFilterButton.setStyleSheet("""
            QPushButton { 
                background-color: #95a5a6; 
                color: white; 
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
            QPushButton:pressed {
                background-color: #707b7c;
            }
        """)

        self.filter_toolbar = QToolBar("Фильтры")
        self.filter_toolbar.setStyleSheet("""
            QToolBar {
                background-color: #34495e;
                spacing: 5px;
                padding: 5px;
            }
        """)
        self.filter_toolbar.addWidget(self.filterButton)
        self.filter_toolbar.addWidget(self.resetFilterButton)
        self.addToolBar(self.filter_toolbar)

    def update_button_visibility(self):
        """Обновляет видимость кнопок в зависимости от текущей таблицы"""
        is_joined_table = self.current_table == 'joined_experts'
        is_group_table = self.current_table == 'expert_group'

        self.filter_toolbar.setVisible(is_joined_table)
        self.addButton.setVisible(not is_joined_table and not is_group_table)
        self.editButton.setVisible(not is_joined_table)
        self.deleteButton.setVisible(not is_joined_table)

        # Кнопка просмотра состава видна только для таблицы групп
        self.viewMembersButton.setVisible(is_group_table)

        # Обновляем состояние кнопки просмотра состава
        if is_group_table:
            has_selection = self.table_widget.currentRow() != -1
            self.viewMembersButton.setEnabled(has_selection)

    def setup_table_behavior(self):
        """Настройка поведения таблицы"""
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_widget.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_widget.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)  # Множественное выделение
        self.table_widget.setAlternatingRowColors(True)

        # Подключаем сигнал изменения выделения для обновления состояния кнопки просмотра состава
        self.table_widget.selectionModel().selectionChanged.connect(self.on_selection_changed)

    def on_selection_changed(self):
        """Обработчик изменения выделения в таблице"""
        if self.current_table == 'expert_group':
            has_selection = self.table_widget.currentRow() != -1
            self.viewMembersButton.setEnabled(has_selection)

    def format_date(self, date_string):
        """Преобразование даты в формат для отображения"""
        if not date_string:
            return ""
        return DateValidator.format_date_for_display(date_string) or str(date_string)

    def format_grnti_code(self, code):
        """Форматирует код ГРНТИ, делая первую часть двузначной"""
        parts = str(code).split('.')
        if len(parts) > 0:
            # Делаем первую часть двузначной
            parts[0] = parts[0].zfill(2)
        return '.'.join(parts)

    def format_grnti_display(self, value):
        """Форматирует значение, которое может содержать один или несколько кодов ГРНТИ"""
        if not value:
            return ""

        if ';' in value:
            # Несколько кодов
            codes = value.split(';')
            formatted_codes = []
            for code in codes:
                formatted_codes.append(self.format_grnti_code(code.strip()))
            return '; '.join(formatted_codes)
        else:
            return self.format_grnti_code(value)

    def connect_menu_actions(self):
        """Связываем пункты меню с соответствующими таблицами"""
        self.actionExperts.triggered.connect(lambda: self.show_table("expert"))
        self.actionGRNTI.triggered.connect(lambda: self.show_table("grnti_classifier"))
        self.actionRegions.triggered.connect(lambda: self.show_table("reg_obl_city"))

        self.actionJoined = QAction("Общая таблица", self)
        self.menu.addAction(self.actionJoined)
        self.actionJoined.triggered.connect(lambda: self.show_table("joined_experts"))

        self.actionGroups = QAction("Группы", self)
        self.menu_2.addAction(self.actionGroups)
        self.actionGroups.triggered.connect(lambda: self.show_table("expert_group"))

    def connect_button_actions(self):
        """Связываем кнопки с функциями"""
        self.addButton.clicked.connect(self.add_record)
        self.editButton.clicked.connect(self.edit_record)
        self.deleteButton.clicked.connect(self.delete_record)

        button_style = """
            QPushButton {
                background-color: #2c3e50;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 4px;
                font-weight: bold;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #34495e;
            }
            QPushButton:pressed {
                background-color: #1a252f;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
                color: #bdc3c7;
            }
        """

        self.addButton.setStyleSheet(button_style)
        self.editButton.setStyleSheet(button_style)
        self.deleteButton.setStyleSheet(button_style)

    def add_experts_to_group(self, expert_ids, expert_names):
        """Добавляет выбранных экспертов в группу - ОБНОВЛЕННАЯ ВЕРСИЯ"""
        try:
            if not expert_ids:
                QMessageBox.warning(self, "Ошибка", "Не выбраны эксперты для добавления")
                return

            # Проверяем, есть ли вообще группы в системе
            groups_exist = len(self.db.get_table_data('expert_group') or []) > 0

            # Создаем диалог с учетом наличия групп
            dialog = AddToGroupDialog(self, self.db, expert_ids, expert_names)

            # Если групп нет, автоматически выбираем создание новой группы
            if not groups_exist:
                # Находим индекс опции "Создать новую группу"
                for i in range(dialog.action_combo.count()):
                    if dialog.action_combo.itemData(i) == "new":
                        dialog.action_combo.setCurrentIndex(i)
                        break

            if dialog.exec():
                action_data = dialog.get_selected_action()

                if action_data['action'] == 'existing':
                    # Добавляем в существующую группу
                    group_id = action_data['group_id']
                    added_count = self.db.add_experts_to_group(group_id, expert_ids)

                    if added_count > 0:
                        # ОБНОВЛЯЕМ поля регионов и рубрик для существующей группы
                        self.update_group_fields(group_id)

                        QMessageBox.information(self, "Успех",
                                                f"Успешно добавлено {added_count} экспертов в группу: {action_data['group_name']}")

                        # Показываем сообщение о дубликатах, если есть
                        if added_count < len(expert_ids):
                            duplicate_count = len(expert_ids) - added_count
                            QMessageBox.information(self, "Информация",
                                                    f"{duplicate_count} экспертов уже находились в группе и не были добавлены повторно")
                    else:
                        QMessageBox.information(self, "Информация",
                                                "Все выбранные эксперты уже находятся в этой группе")

                else:
                    # Создаем новую группу
                    group_name = action_data['group_name']
                    group_id = self.db.create_group_with_experts(group_name, expert_ids)

                    # ОБНОВЛЯЕМ поля регионов и рубрик для новой группы
                    self.update_group_fields(group_id)

                    QMessageBox.information(self, "Успех",
                                            f"Создана новая группа '{group_name}' с {len(expert_ids)} экспертами")

                # Обновляем таблицу групп, если она открыта
                if self.current_table == 'expert_group':
                    self.show_table('expert_group')

        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось добавить экспертов в группу: {str(e)}")

    def open_filter_dialog(self):
        """Открывает диалог фильтрации"""
        if self.current_table != 'joined_experts':
            return

        dialog = FilterDialog(self, self.db, self.current_filters)
        if dialog.exec():
            new_filters = dialog.get_filters()
            if new_filters:
                self.current_filters = new_filters
                self.apply_filters()
            else:
                # Если фильтры пустые, сбрасываем их
                self.current_filters = {}
                self.show_table('joined_experts')
                self.statusbar.showMessage("Фильтры сброшены. Отображены все эксперты.")

    def apply_filters(self):
        """Применяет текущие фильтры к общей таблице"""
        if self.current_table == 'joined_experts':
            self.show_table('joined_experts',
                            self.current_sort_column,
                            'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC')

            filter_info = self.get_filter_info()
            current_message = self.statusbar.currentMessage()
            base_message = current_message.split('|')[0] if '|' in current_message else current_message
            self.statusbar.showMessage(f"{base_message} | {filter_info}")

    def get_filter_info(self):
        """Возвращает строку с информацией о примененных фильтрах"""
        if not self.current_filters:
            return "Фильтры не применены"

        filter_parts = []
        display_names = {
            'expert_names': 'ФИО',
            'regions': 'Регион',
            'cities': 'Город',
            'keywords': 'Ключ. слова',
            'group_conditions': 'Кол-во групп',
            'grnti_codes': 'Коды ГРНТИ',
            'subrubrics': 'Подрубрики',
            'disciplines': 'Дисциплины'
        }

        for key, value in self.current_filters.items():
            if key in display_names:
                if isinstance(value, list):
                    if key == 'group_conditions':
                        filter_parts.append(f"{display_names[key]}: {', '.join(map(str, value))}")
                    else:
                        filter_parts.append(f"{display_names[key]}: {', '.join(map(str, value))}")
                else:
                    filter_parts.append(f"{display_names[key]}: {value}")

        return f"Фильтры: {', '.join(filter_parts)}"

    def reset_filters(self):
        """Сбрасывает все фильтры"""
        self.current_filters = {}
        if self.current_table == 'joined_experts':
            self.show_table('joined_experts')
            self.statusbar.showMessage("Фильтры сброшены. Отображены все эксперты.")

    def show_table(self, table_name, sort_column=None, sort_order='ASC'):
        """Отображение содержимого таблицы с возможностью сортировки и фильтрации"""
        try:
            self.current_table = table_name

            if sort_column is None and table_name in self.sort_settings:
                sort_column, sort_order = self.sort_settings[table_name]

            self.update_button_visibility()

            # Используем специальный метод для таблицы экспертов с ГРНТИ
            if table_name == 'expert':
                data = self.db.get_experts_with_grnti(sort_column, sort_order)
                columns = ['id', 'name', 'region', 'city', 'input_date', 'keywords', 'group_count', 'grnti']
            elif table_name == 'joined_experts':
                if self.current_filters:
                    data = self.db.get_filtered_joined_experts_data(self.current_filters, sort_column, sort_order)
                else:
                    data = self.db.get_joined_experts_data(sort_column, sort_order)
                # Для joined_experts SQL возвращает: expert_id, expert_name, region, city, input_date, keywords, group_count, grnti, grnti_descriptions
                # Но expert_id скрывается, поэтому создаем правильное сопоставление
                columns = ['expert_id', 'expert_name', 'region', 'city', 'input_date', 'keywords', 'group_count',
                           'grnti', 'grnti_descriptions']
            else:
                data = self.db.get_table_data(table_name, sort_column, sort_order)
                columns = self.db.get_columns_names(table_name)

            self.current_table_data[table_name] = data

            self.table_widget.clear()
            self.table_widget.setSortingEnabled(False)

            display_columns = []
            for col in columns:
                if self.should_hide_column(table_name, col):
                    continue
                if (table_name in self.column_display_names and
                        col in self.column_display_names[table_name]):
                    display_columns.append(self.column_display_names[table_name][col])
                else:
                    display_columns.append(col)

            self.table_widget.setColumnCount(len(display_columns))

            header_labels = []
            for i, col in enumerate(display_columns):
                db_column_name = self.get_db_column_name(table_name, i)

                if sort_column and db_column_name == sort_column:
                    arrow = " ↑" if sort_order == 'ASC' else " ↓"
                    header_labels.append(f"{col}{arrow}")
                else:
                    header_labels.append(col)

            self.table_widget.setHorizontalHeaderLabels(header_labels)
            self.table_widget.setRowCount(len(data))

            date_columns = self.date_columns.get(table_name, [])

            for row_num, row_data in enumerate(data):
                col_num_display = 0

                for col_num, value in enumerate(row_data):
                    col_name = columns[col_num]

                    if self.should_hide_column(table_name, col_name):
                        continue

                    # Форматирование кода ГРНТИ, если необходимо
                    if ((table_name == 'grnti_classifier' and col_name == 'codrub') or
                            (table_name == 'expert_grnti' and col_name == 'rubric') or
                            (table_name == 'expert' and col_name == 'grnti') or
                            (table_name == 'joined_experts' and col_name == 'grnti')):
                        value = self.format_grnti_display(str(value))
                    elif col_name in date_columns and value is not None:
                        if hasattr(value, 'strftime'):
                            value = value.strftime('%d.%m.%Y')
                        else:
                            value = self.format_date(str(value))
                    else:
                        value = str(value) if value is not None else ""

                    item = QTableWidgetItem(value)
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)

                    if table_name == 'grnti_classifier' and col_name == 'codrub':
                        try:
                            numeric_value = int(value) if value is not None else 0
                            item.setData(Qt.ItemDataRole.UserRole, numeric_value)
                        except (ValueError, TypeError):
                            item.setData(Qt.ItemDataRole.UserRole, 0)
                    elif table_name == 'joined_experts':
                        if col_name == 'group_count':
                            try:
                                numeric_value = int(value) if value is not None else 0
                                item.setData(Qt.ItemDataRole.UserRole, numeric_value)
                            except (ValueError, TypeError):
                                item.setData(Qt.ItemDataRole.UserRole, 0)

                    self.table_widget.setItem(row_num, col_num_display, item)
                    col_num_display += 1

            self.table_widget.resizeColumnsToContents()
            header = self.table_widget.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

            self.table_widget.clearSelection()
            self.table_widget.setAlternatingRowColors(True)
            self.table_widget.resizeRowsToContents()

            # УЛУЧШЕННАЯ АВТОПРОКРУТКА К НОВОЙ ЗАПИСИ
            if hasattr(self, 'last_added_id') and self.last_added_id:
                # Используем таймер для гарантированного выделения после отображения таблицы
                QTimer.singleShot(150, lambda: self.highlight_and_scroll_to_record(table_name, self.last_added_id))

            sort_status = ""
            if sort_column:
                sort_direction = "↑" if sort_order == 'ASC' else "↓"
                display_name = self.column_display_names.get(table_name, {}).get(sort_column, sort_column)
                sort_status = f" | Сортировка: {display_name} {sort_direction}"

            filter_status = self.get_filter_info() if table_name == 'joined_experts' and self.current_filters else ""

            self.statusbar.showMessage(f"Таблица: {table_name} | Записей: {len(data)}{sort_status} | {filter_status}")

        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось загрузить таблицу: {str(e)}")

    def highlight_and_scroll_to_record(self, table_name, record_id):
        """Выделяет и прокручивает к добавленной записи в таблице - ИСПРАВЛЕННАЯ ВЕРСИЯ"""
        print(f"Поиск записи с ID: {record_id} в таблице {table_name}")

        if table_name == 'expert':
            # Для таблицы экспертов ищем по ID в исходных данных
            # В таблице expert первый столбец скрыт (id), поэтому ищем во всех данных
            for row in range(len(self.current_table_data.get(table_name, []))):
                if self.current_table_data[table_name][row][0] == record_id:
                    # В таблице expert первый столбец (ID) скрыт, поэтому выделяем строку
                    self.table_widget.selectRow(row)
                    self.table_widget.scrollToItem(self.table_widget.item(row, 0))
                    print(f"Найдена запись в строке {row}")
                    # Сбрасываем last_added_id после успешного поиска
                    self.last_added_id = None
                    return
        elif table_name == 'joined_experts':
            # Для общей таблицы ищем по первому столбцу (expert_id)
            for row in range(self.table_widget.rowCount()):
                item = self.table_widget.item(row, 0)  # Первый столбец - expert_id
                if item and item.text().isdigit() and int(item.text()) == record_id:
                    self.table_widget.selectRow(row)
                    self.table_widget.scrollToItem(item)
                    print(f"Найдена запись в строке {row}")
                    # Сбрасываем last_added_id после успешного поиска
                    self.last_added_id = None
                    return
        else:
            # Для других таблиц ищем в первом столбце
            for row in range(self.table_widget.rowCount()):
                item = self.table_widget.item(row, 0)
                if item and item.text().isdigit() and int(item.text()) == record_id:
                    self.table_widget.selectRow(row)
                    self.table_widget.scrollToItem(item)
                    print(f"Найдена запись в строке {row}")
                    # Сбрасываем last_added_id после успешного поиска
                    self.last_added_id = None
                    return

        print("Запись не найдена")

    def add_record(self):
        """Добавить новую запись"""
        if not self.current_table:
            QMessageBox.warning(self, "Ошибка", "Сначала выберите таблицу")
            return

        if self.current_table == 'joined_experts':
            QMessageBox.information(self, "Информация",
                                    "Добавление записей в общей таблице невозможно.\n"
                                    "Используйте соответствующие таблицы экспертов и кодов ГРНТИ.")
            return

        if self.current_table == 'expert_group':
            QMessageBox.information(self, "Информация",
                                    "Создание групп доступно только из вкладки 'Общая таблица'.")
            return

        try:
            columns = self.db.get_columns_names(self.current_table)
            display_names = self.column_display_names.get(self.current_table, {})
            date_columns = self.date_columns.get(self.current_table, [])
            numeric_columns = self.numeric_columns.get(self.current_table, [])

            if self.current_table == 'expert':
                columns_without_id = columns[1:]
                display_names_without_id = {k: v for k, v in display_names.items() if k != 'id'}

                dialog = ExpertEditDialog(
                    self.current_table,
                    columns_without_id,
                    data=None,
                    parent=self,
                    db=self.db,
                    is_edit=False,
                    expert_id=None
                )

                if dialog.exec():
                    data_without_id = dialog.get_data()
                    grnti_data = dialog.get_grnti_data()
                    next_id = self.get_next_expert_id()

                    full_data = [next_id] + data_without_id

                    self.db.insert_expert_record(full_data)

                    if grnti_data:
                        self.db.save_expert_grnti(next_id, grnti_data)

                    # Сохраняем ID для выделения
                    self.last_added_id = next_id
                    expert_name = data_without_id[0] if data_without_id and data_without_id[0] else None
                    if expert_name:
                        self.statusbar.showMessage(f"Эксперт '{expert_name}' успешно добавлен")
                    else:
                        self.statusbar.showMessage("Эксперт успешно добавлен")

                    # Немедленное обновление таблицы и прокрутка
                    self.show_table(self.current_table, self.current_sort_column,
                                    'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC')

            elif self.current_table == 'expert_group':
                dialog = GroupEditDialog(parent=self, db=self.db, is_edit=False)
                if dialog.exec():
                    data = dialog.get_data()
                    print(f"Создание группы: {data['name']}")
                    print(f"Эксперты: {[expert_id for expert_id, _ in data['experts']]}")
                    print(f"Регионы: {data['regions']}")
                    print(f"Рубрики: {data['rubric_code']}")

                    # Создаем группу с экспертами через специальный метод
                    expert_ids = [expert_id for expert_id, _ in data['experts']]

                    # РАЗРЕШАЕМ СОЗДАНИЕ ПУСТОЙ ГРУППЫ БЕЗ ЭКСПЕРТОВ
                    try:
                        group_id = self.db.create_group_with_experts(data['name'], expert_ids)

                        if group_id:
                            # Обновляем дополнительные поля группы (регионы и рубрики)
                            cursor = self.db.connection.cursor()
                            try:
                                update_query = """
                                    UPDATE expert_group 
                                    SET regions = %s, rubric_code = %s 
                                    WHERE id = %s
                                """
                                print(
                                    f"Обновление группы {group_id}: regions='{data['regions']}', rubric_code='{data['rubric_code']}'")

                                cursor.execute(update_query, (data['regions'], data['rubric_code'], group_id))
                                self.db.connection.commit()
                                print("Дополнительные поля группы успешно обновлены")

                            except Exception as e:
                                self.db.connection.rollback()
                                print(f"Ошибка обновления дополнительных полей группы: {e}")
                                # Создаем сообщение об ошибке, но не прерываем выполнение
                                QMessageBox.warning(self, "Предупреждение",
                                                    f"Группа создана, но не удалось сохранить дополнительные поля: {str(e)}")
                            finally:
                                cursor.close()

                        self.last_added_id = group_id
                        self.statusbar.showMessage(
                            f"Группа '{data['name']}' успешно создана с {len(expert_ids)} экспертами")
                        self.show_table(self.current_table)

                    except Exception as e:
                        QMessageBox.critical(self, "Ошибка", f"Не удалось создать группу: {str(e)}")

            elif self.current_table == 'grnti_classifier':
                display_names['_table_name'] = self.table_display_names.get(self.current_table, self.current_table)
                dialog = EditDialog(
                    self.current_table,
                    columns,
                    data=None,
                    parent=self,
                    display_names=display_names,
                    date_columns=date_columns,
                    is_edit=False,
                    numeric_columns=numeric_columns
                )

                if dialog.exec():
                    data = dialog.get_data()
                    try:
                        # Для ГРНТИ преобразуем код в число
                        if data[0]:  # codrub поле
                            try:
                                data[0] = int(data[0])
                            except ValueError:
                                QMessageBox.warning(self, "Ошибка", "Код ГРНТИ должен быть числом")
                                return

                        inserted_id = self.db.insert_record(self.current_table, data)
                        if inserted_id:
                            self.last_added_id = inserted_id
                        self.statusbar.showMessage("Код ГРНТИ успешно добавлен")
                        self.show_table(self.current_table, self.current_sort_column,
                                        'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC')
                    except Exception as e:
                        QMessageBox.warning(self, "Ошибка", f"Не удалось добавить код ГРНТИ: {str(e)}")

            elif self.current_table == 'reg_obl_city':
                display_names['_table_name'] = self.table_display_names.get(self.current_table, self.current_table)
                dialog = EditDialog(
                    self.current_table,
                    columns,
                    data=None,
                    parent=self,
                    display_names=display_names,
                    date_columns=date_columns,
                    is_edit=False,
                    numeric_columns=numeric_columns
                )

                if dialog.exec():
                    data = dialog.get_data()
                    try:
                        if columns and columns[0].lower() == 'id':
                            self.db.insert_record_without_id(self.current_table, data[1:])
                            self.last_added_id = None
                        else:
                            inserted_id = self.db.insert_record(self.current_table, data)
                            if inserted_id:
                                self.last_added_id = inserted_id
                        self.statusbar.showMessage("Регион/город успешно добавлен")
                        self.show_table(self.current_table, self.current_sort_column,
                                        'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC')
                    except Exception as e:
                        QMessageBox.warning(self, "Ошибка", f"Не удалось добавить регион/город: {str(e)}")

            elif self.current_table == 'expert_grnti':
                QMessageBox.information(
                    self,
                    "Информация",
                    "Связи экспертов с ГРНТИ можно редактировать из карточки эксперта."
                )
                return

            else:
                # Универсальный обработчик для других таблиц
                display_names['_table_name'] = self.table_display_names.get(self.current_table, self.current_table)
                dialog = EditDialog(
                    self.current_table,
                    columns,
                    data=None,
                    parent=self,
                    display_names=display_names,
                    date_columns=date_columns,
                    is_edit=False,
                    numeric_columns=numeric_columns
                )

                if dialog.exec():
                    data = dialog.get_data()
                    try:
                        if columns and columns[0].lower() == 'id':
                            self.db.insert_record_without_id(self.current_table, data[1:])
                            self.last_added_id = None
                        else:
                            inserted_id = self.db.insert_record(self.current_table, data)
                            if inserted_id:
                                self.last_added_id = inserted_id
                        self.statusbar.showMessage("Запись успешно добавлена")
                        self.show_table(self.current_table, self.current_sort_column,
                                        'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC')
                    except Exception as e:
                        QMessageBox.warning(self, "Ошибка", f"Не удалось добавить запись: {str(e)}")

        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось добавить запись: {str(e)}")

    def edit_record(self):
        """Редактировать выбранную запись"""
        if not self.current_table:
            QMessageBox.warning(self, "Ошибка", "Сначала выберите таблицу")
            return
        if self.current_table == 'joined_experts':
            QMessageBox.information(self, "Информация",
                                    "Редактирование записей в общей таблице невозможно.\n"
                                    "Используйте соответствующие таблицы экспертов и кодов ГРНТИ.")
            return
        if self.current_table == 'expert_grnti':
            QMessageBox.information(
                self,
                "Информация",
                "Связи экспертов с ГРНТИ можно редактировать из карточки эксперта."
            )
            return
        selected_row = self.table_widget.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите запись для редактирования")
            return

        # Получаем ID записи
        record_id = None
        if self.current_table in self.current_table_data:
            if selected_row < len(self.current_table_data[self.current_table]):
                record_id = self.current_table_data[self.current_table][selected_row][0]
        else:
            id_item = self.table_widget.item(selected_row, 0)
            if id_item and id_item.text().isdigit():
                record_id = int(id_item.text())

        if not record_id:
            QMessageBox.warning(self, "Ошибка", "Не удалось определить запись")
            return

        try:
            # Получаем актуальные данные записи из БД
            table_data = self.db.get_table_data(
                self.current_table,
                self.current_sort_column,
                'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC'
            )
            raw_row_data = None
            for row in table_data:
                if row[0] == record_id:
                    raw_row_data = row
                    break
            if not raw_row_data:
                QMessageBox.warning(self, "Ошибка", "Не удалось найти запись для редактирования")
                return

            # === ОБРАБОТКА РАЗНЫХ ТИПОВ ТАБЛИЦ ===
            if self.current_table == 'expert':
                columns = self.db.get_columns_names(self.current_table)
                raw_data_without_id = list(raw_row_data[1:])
                columns_without_id = columns[1:]
                dialog = ExpertEditDialog(
                    self.current_table,
                    columns_without_id,
                    raw_data_without_id,
                    parent=self,
                    db=self.db,
                    is_edit=True,
                    expert_id=record_id
                )
                if dialog.exec():
                    data = dialog.get_data()
                    self.db.update_expert_record(record_id, data)
                    grnti_data = dialog.get_grnti_data()
                    if grnti_data:
                        self.db.save_expert_grnti(record_id, grnti_data)
                    self.show_table(self.current_table, self.current_sort_column,
                                    'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC')
                    self.statusbar.showMessage("Запись успешно обновлена")

            elif self.current_table == 'expert_group':
                # Проверяем, что данных достаточно (6 полей)
                if len(raw_row_data) < 6:
                    QMessageBox.warning(self, "Ошибка", "Неполные данные группы")
                    return

                # Для групп показываем диалог редактирования
                dialog = GroupEditDialog(
                    parent=self,
                    db=self.db,
                    is_edit=True,
                    group_data=raw_row_data,
                    group_id=record_id
                )
                if dialog.exec():
                    data = dialog.get_data()

                    # Получаем текущих участников группы
                    current_members = self.db.get_group_members(record_id)
                    current_expert_ids = [member[0] for member in current_members]
                    new_expert_ids = [expert_id for expert_id, _ in data['experts']]

                    # Находим экспертов для добавления и удаления
                    experts_to_add = [expert_id for expert_id in new_expert_ids if expert_id not in current_expert_ids]
                    experts_to_remove = [expert_id for expert_id in current_expert_ids if
                                         expert_id not in new_expert_ids]

                    # Удаляем экспертов из группы с ОБНОВЛЕНИЕМ СЧЕТЧИКОВ
                    if experts_to_remove:
                        removed_count = self.db.remove_experts_from_group(record_id, experts_to_remove)
                        print(f"Удалено {removed_count} экспертов из группы {record_id}")

                    # Добавляем новых экспертов в группу с ОБНОВЛЕНИЕМ СЧЕТЧИКОВ
                    if experts_to_add:
                        added_count = self.db.add_experts_to_group(record_id, experts_to_add)
                        print(f"Добавлено {added_count} экспертов в группу {record_id}")

                    # Обновляем данные группы
                    update_data = [
                        data['name'],
                        data['participant_count'],
                        data['created_at'],
                        data['regions'],
                        data['rubric_code']
                    ]
                    self.db.update_record('expert_group', record_id, update_data)

                    self.statusbar.showMessage("Группа успешно обновлена")
                    self.show_table(self.current_table)

            else:
                # Универсальный редактор для остальных таблиц
                columns = self.db.get_columns_names(self.current_table)
                display_names = self.column_display_names.get(self.current_table, {})
                date_columns = self.date_columns.get(self.current_table, [])
                numeric_columns = self.numeric_columns.get(self.current_table, [])
                row_data = []
                for i, value in enumerate(raw_row_data):
                    if value is None:
                        row_data.append("")
                    elif columns[i] in date_columns and hasattr(value, 'strftime'):
                        formatted_date = DateValidator.format_date_for_display(value.strftime('%Y-%m-%d'))
                        row_data.append(formatted_date if formatted_date else str(value))
                    else:
                        row_data.append(str(value))
                display_names['_table_name'] = self.table_display_names.get(self.current_table, self.current_table)
                dialog = EditDialog(
                    self.current_table,
                    columns,
                    row_data,
                    parent=self,
                    display_names=display_names,
                    date_columns=date_columns,
                    is_edit=True,
                    numeric_columns=numeric_columns
                )
                if dialog.exec():
                    data = dialog.get_data()
                    self.db.update_record(self.current_table, record_id, data[1:])
                    self.show_table(self.current_table, self.current_sort_column,
                                    'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC')
                    self.statusbar.showMessage("Запись успешно обновлена")

        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось обновить запись: {str(e)}")

    def delete_record(self):
        """Удалить выбранную запись - С ПРОВЕРКОЙ ГРУПП ДЛЯ ЭКСПЕРТОВ"""
        if not self.current_table:
            QMessageBox.warning(self, "Ошибка", "Сначала выберите таблицу")
            return

        if self.current_table == 'joined_experts':
            QMessageBox.information(self, "Информация",
                                    "Удаление записей из общей таблицы невозможно.\n"
                                    "Используйте соответствующие таблицы экспертов и кодов ГРНТИ.")
            return

        selected_row = self.table_widget.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите запись для удаления")
            return

        try:
            # Получаем актуальные данные таблицы с текущей сортировкой
            if self.current_table == 'expert':
                table_data = self.db.get_experts_with_grnti(
                    self.current_sort_column,
                    'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC'
                )
            elif self.current_table == 'joined_experts':
                if self.current_filters:
                    table_data = self.db.get_filtered_joined_experts_data(
                        self.current_filters,
                        self.current_sort_column,
                        'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC'
                    )
                else:
                    table_data = self.db.get_joined_experts_data(
                        self.current_sort_column,
                        'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC'
                    )
            else:
                table_data = self.db.get_table_data(
                    self.current_table,
                    self.current_sort_column,
                    'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC'
                )

            if selected_row >= len(table_data):
                QMessageBox.warning(self, "Ошибка", "Неверный индекс строки")
                return

            # Получаем ID записи из актуальных данных
            record_id = table_data[selected_row][0]

            # Получаем имя записи для сообщения
            record_name = ""
            if self.current_table == 'expert' and len(table_data[selected_row]) > 1:
                record_name = f" - {table_data[selected_row][1]}"
            elif self.current_table == 'grnti_classifier' and len(table_data[selected_row]) > 1:
                record_name = f" - {table_data[selected_row][1]}"
            elif self.current_table == 'expert_group' and len(table_data[selected_row]) > 1:
                record_name = f" - {table_data[selected_row][1]}"

            # Проверка зависимостей для таблицы экспертов
            if self.current_table == 'expert':
                has_dependencies, dependency_info = self.db.check_expert_dependencies(record_id)
                if has_dependencies:
                    QMessageBox.warning(
                        self,
                        "Ошибка удаления",
                        f"Невозможно удалить эксперта, так как он состоит в группах.\n\n"
                        f"{dependency_info}\n\n"
                        f"Пожалуйста, сначала удалите эксперта из всех групп, а затем повторите удаление."
                    )
                    return

            # Проверка зависимостей для таблицы ГРНТИ
            if self.current_table == 'grnti_classifier':
                has_dependencies = self.db.check_grnti_dependencies(record_id)
                if has_dependencies:
                    QMessageBox.warning(
                        self,
                        "Ошибка удаления",
                        "Невозможно удалить запись ГРНТИ, так как она используется в таблице экспертов.\n\n"
                        "Пожалуйста, сначала удалите все ссылки на этот код ГРНТИ в таблице экспертов."
                    )
                    return

            # Проверка зависимостей для таблицы регионов
            if self.current_table == 'reg_obl_city':
                has_dependencies, dependency_info = self.db.check_region_city_dependencies(record_id)
                if has_dependencies:
                    QMessageBox.warning(
                        self,
                        "Ошибка удаления",
                        f"Невозможно удалить запись региона/города, так как она используется в таблице экспертов.\n\n"
                        f"Зависимости: {dependency_info}\n\n"
                        f"Пожалуйста, сначала удалите или измените все ссылки на этот регион/город в таблице экспертов."
                    )
                    return

            # Создаем кастомное сообщение с русскими кнопками
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Подтверждение удаления")
            msg_box.setText(f"Вы уверены, что хотите удалить запись: {record_id}{record_name}?")
            msg_box.setIcon(QMessageBox.Icon.Question)

            # Создаем кнопки с русским текстом
            yes_button = msg_box.addButton("Да", QMessageBox.ButtonRole.YesRole)
            no_button = msg_box.addButton("Нет", QMessageBox.ButtonRole.NoRole)
            msg_box.setDefaultButton(no_button)

            # Стилизация кнопок
            yes_button.setStyleSheet("""
                QPushButton {
                    background-color: #e74c3c;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-weight: bold;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            no_button.setStyleSheet("""
                QPushButton {
                    background-color: #95a5a6;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-weight: bold;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #7f8c8d;
                }
            """)

            msg_box.exec()

            if msg_box.clickedButton() == yes_button:
                if self.current_table == 'expert':
                    # Для экспертов сначала удаляем связанные коды ГРНТИ
                    self.db.save_expert_grnti(record_id, [])

                if self.current_table == 'expert_group':
                    # Для групп используем специальную функцию, которая обновляет счетчики групп у экспертов
                    self.db.delete_group(record_id)
                else:
                    self.db.delete_record(self.current_table, record_id)
                self.show_table(self.current_table, self.current_sort_column,
                                'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC')
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось удалить запись: {str(e)}")

    def show_context_menu(self, position):
        """Показывает контекстное меню - ВЕРСИЯ С ИСПОЛЬЗОВАНИЕМ ИСХОДНЫХ ДАННЫХ"""
        if self.current_table != 'joined_experts':
            return

        print("Контекстное меню вызвано для joined_experts")

        # Получаем выбранные строки
        selected_rows = self.table_widget.selectionModel().selectedRows()
        print(f"Выбранных строк: {len(selected_rows)}")

        if not selected_rows:
            index = self.table_widget.indexAt(position)
            if index.isValid():
                self.table_widget.selectRow(index.row())
                selected_rows = self.table_widget.selectionModel().selectedRows()
                print(f"После выделения: {len(selected_rows)}")

        if not selected_rows:
            return

        context_menu = QMenu(self)
        expert_ids = []
        expert_names = []

        # Получаем исходные данные таблицы
        try:
            if self.current_filters:
                table_data = self.db.get_filtered_joined_experts_data(
                    self.current_filters,
                    self.current_sort_column,
                    'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC'
                )
            else:
                table_data = self.db.get_joined_experts_data(
                    self.current_sort_column,
                    'ASC' if self.current_sort_order == Qt.SortOrder.AscendingOrder else 'DESC'
                )

            print(f"Всего данных в таблице: {len(table_data)}")

            for model_index in selected_rows:
                row = model_index.row()
                print(f"Обрабатываем строку {row}")

                if row < len(table_data):
                    row_data = table_data[row]
                    print(f"Данные строки: {row_data}")

                    # В joined_experts структура: (expert_id, expert_name, region, city, input_date, keywords, group_count, grnti, grnti_descriptions)
                    if len(row_data) >= 2:
                        expert_id = row_data[0]  # Первый элемент - expert_id
                        expert_name = row_data[1]  # Второй элемент - expert_name

                        if expert_id is not None:
                            expert_ids.append(expert_id)
                            expert_names.append(expert_name if expert_name else f"Эксперт {expert_id}")
                            print(f"Добавлен из исходных данных: ID={expert_id}, ФИО={expert_name}")
                        else:
                            print("ID эксперта None")
                    else:
                        print(f"Недостаточно данных в строке: {len(row_data)}")
                else:
                    print(f"Строка {row} выходит за пределы данных")

        except Exception as e:
            print(f"Ошибка при получении данных: {e}")

        print(f"Итоговый список: {len(expert_ids)} экспертов")

        if expert_ids:
            action_text = f"Добавить {len(expert_ids)} экспертов в группу"
            add_to_group_action = context_menu.addAction(action_text)
            add_to_group_action.triggered.connect(lambda: self.add_experts_to_group(expert_ids, expert_names))

            info_action = context_menu.addAction(f"Выбрано экспертов: {len(expert_ids)}")
            info_action.setEnabled(False)
        else:
            no_experts_action = context_menu.addAction("Не удалось получить данные экспертов")
            no_experts_action.setEnabled(False)

        context_menu.exec(self.table_widget.viewport().mapToGlobal(position))

    def setup_visual_feedback(self):
        """Настройка визуальной обратной связи для таблицы"""
        self.table_widget.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #f8f9fa;
                gridline-color: #dee2e6;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                color: black;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #dee2e6;
                color: black;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QHeaderView::section {
                background-color: #2c3e50;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
            }
            QHeaderView::section:checked {
                background-color: #34495e;
            }
        """)

    def get_next_expert_id(self):
        """Получить следующий доступный ID для эксперта"""
        try:
            cursor = self.db.connection.cursor()
            cursor.execute('SELECT MAX(id) FROM expert')
            max_id = cursor.fetchone()[0]
            cursor.close()

            if max_id is None:
                return 1
            else:
                return max_id + 1
        except Exception as e:
            return 1

    def closeEvent(self, event):
        """Закрытие соединения с базой данных при выходе"""
        if hasattr(self, 'db'):
            self.db.connection.close()
            print("Соединение с базой данных закрыто")
        event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)

    app.setStyleSheet("""
        QMessageBox {
            background-color: white;
        }
        QMessageBox QLabel {
            color: #2c3e50;
        }
    """)

    window = MainWindow()
    window.show()
    sys.exit(app.exec())